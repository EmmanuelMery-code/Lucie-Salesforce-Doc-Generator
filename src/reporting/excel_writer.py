from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.models import ObjectInfo, PmdViolation, SecurityArtifact


# Excel spec allows a theoretically unbounded number of sheets per workbook but
# the file format becomes unresponsive well before that. We cap at a safe soft
# limit (1 index sheet + N object sheets) and roll over into a "part 2"
# workbook past the threshold.
MAX_OBJECT_SHEETS_PER_WORKBOOK = 200

# Excel forbids certain characters in sheet names and enforces a 31-char limit.
_FORBIDDEN_SHEET_CHARS_RE = re.compile(r"[:\\/?*\[\]]")


class ExcelReportWriter:
    def __init__(self, log_callback=None) -> None:
        self.log = log_callback or (lambda message: None)

    def write_security_workbook(
        self, artifacts: list[SecurityArtifact], output_path: str | Path, title: str
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Synthese"
        self._write_sheet(
            summary,
            [
                "Nom",
                "Label",
                "Description",
                "Nb droits objet",
                "Nb droits champ",
                "Nb permissions systeme",
                "Nb applis",
                "Nb onglets",
                "Nb classes",
                "Nb flows",
            ],
            [
                [
                    artifact.name,
                    artifact.label,
                    artifact.description,
                    len(artifact.object_permissions),
                    len(artifact.field_permissions),
                    len(artifact.user_permissions),
                    len(artifact.application_visibilities),
                    len(artifact.tab_visibilities),
                    len(artifact.class_accesses),
                    len(artifact.flow_accesses),
                ]
                for artifact in artifacts
            ],
        )

        self._write_sheet(
            workbook.create_sheet("DroitsObjet"),
            ["Nom", "Objet", "Lecture", "Creation", "Modification", "Suppression", "ViewAll", "ModifyAll"],
            [
                [
                    artifact.name,
                    permission.object_name,
                    permission.allow_read,
                    permission.allow_create,
                    permission.allow_edit,
                    permission.allow_delete,
                    permission.view_all_records,
                    permission.modify_all_records,
                ]
                for artifact in artifacts
                for permission in artifact.object_permissions
            ],
        )

        self._write_sheet(
            workbook.create_sheet("DroitsChamp"),
            ["Nom", "Champ", "Lecture", "Modification"],
            [
                [artifact.name, permission.field_name, permission.readable, permission.editable]
                for artifact in artifacts
                for permission in artifact.field_permissions
            ],
        )

        self._write_sheet(
            workbook.create_sheet("PermissionsSysteme"),
            ["Nom", "Permission", "Activee"],
            [
                [artifact.name, permission.name, permission.enabled]
                for artifact in artifacts
                for permission in artifact.user_permissions
            ],
        )

        self._write_sheet(
            workbook.create_sheet("Applications"),
            ["Nom", "Application", "Visible", "Defaut"],
            [
                [artifact.name, app.name, app.visible, app.default]
                for artifact in artifacts
                for app in artifact.application_visibilities
            ],
        )

        self._write_sheet(
            workbook.create_sheet("Onglets"),
            ["Nom", "Onglet", "Visibilite", "Defaut"],
            [
                [artifact.name, tab.name, tab.visible, tab.default]
                for artifact in artifacts
                for tab in artifact.tab_visibilities
            ],
        )

        self._write_sheet(
            workbook.create_sheet("ClassesApex"),
            ["Nom", "Classe Apex", "Active"],
            [
                [artifact.name, access.name, access.enabled]
                for artifact in artifacts
                for access in artifact.class_accesses
            ],
        )

        self._write_sheet(
            workbook.create_sheet("Flows"),
            ["Nom", "Flow", "Actif"],
            [
                [artifact.name, access.name, access.enabled]
                for artifact in artifacts
                for access in artifact.flow_accesses
            ],
        )

        self._write_sheet(
            workbook.create_sheet("RecordTypes"),
            ["Nom", "Record Type", "Visible", "Defaut"],
            [
                [artifact.name, item.record_type, item.visible, item.default]
                for artifact in artifacts
                for item in artifact.record_type_visibilities
            ],
        )

        self._write_sheet(
            workbook.create_sheet("CustomPermissions"),
            ["Nom", "Custom Permission", "Activee"],
            [
                [artifact.name, access.name, access.enabled]
                for artifact in artifacts
                for access in artifact.custom_permissions
            ],
        )

        workbook.save(output)
        self.log(f"{title} genere: {output}")
        return output

    def write_inventory_workbook(
        self, inventory: dict[str, list[dict[str, object]]], output_path: str | Path
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet_definitions = [
            ("RecordTypes", "record_types"),
            ("Layouts", "layouts"),
            ("LightningPages", "lightning_pages"),
            ("ValidationRules", "validation_rules"),
            ("OmniStudio", "omnistudio"),
            ("BusinessRules", "business_rules_engine"),
            ("Flows", "flows"),
            ("PermissionSets", "permission_sets"),
            ("Profiles", "profiles"),
            ("Reports", "reports"),
            ("Dashboards", "dashboards"),
        ]

        first_title, first_key = sheet_definitions[0]
        first_rows = inventory.get(first_key, [])
        summary = workbook.active
        summary.title = first_title
        self._write_dict_sheet(summary, first_rows)

        for title, key in sheet_definitions[1:]:
            self._write_dict_sheet(workbook.create_sheet(title), inventory.get(key, []))

        workbook.save(output)
        self.log(f"Classeur inventaire metadata genere: {output}")
        return output

    def write_pmd_workbook(
        self,
        violations_by_artifact: dict[str, list[PmdViolation]],
        output_path: str | Path,
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Synthese"
        self._write_sheet(
            summary,
            ["Composant", "Violations PMD"],
            [
                [artifact_name, len(violations)]
                for artifact_name, violations in sorted(
                    violations_by_artifact.items(), key=lambda item: item[0].lower()
                )
            ],
        )

        detail_rows = [
            [
                artifact_name,
                violation.rule,
                violation.ruleset,
                violation.priority,
                violation.begin_line,
                violation.end_line,
                violation.message,
                str(violation.file_path),
            ]
            for artifact_name, violations in sorted(
                violations_by_artifact.items(), key=lambda item: item[0].lower()
            )
            for violation in violations
        ]
        self._write_sheet(
            workbook.create_sheet("Violations"),
            ["Composant", "Regle", "Ruleset", "Priorite", "LigneDebut", "LigneFin", "Message", "Fichier"],
            detail_rows,
        )

        workbook.save(output)
        self.log(f"Classeur PMD genere: {output}")
        return output

    def write_data_dictionary_workbooks(
        self,
        objects: list[ObjectInfo],
        output_dir: str | Path,
        *,
        max_object_sheets: int = MAX_OBJECT_SHEETS_PER_WORKBOOK,
    ) -> list[Path]:
        """Generate the Data Dictionary workbook(s).

        Each workbook starts with a "Synthese" sheet listing the objects it
        contains (general info) followed by one sheet per object describing
        its fields. When the number of object sheets exceeds
        ``max_object_sheets`` a new workbook is created (``data_dictionary_part_2.xlsx``,
        ``..._part_3.xlsx`` and so on) so Excel stays responsive.

        Returns the list of written file paths in order.
        """
        output_base = Path(output_dir)
        output_base.mkdir(parents=True, exist_ok=True)

        if not objects:
            # Still produce an (almost) empty workbook so that the index page
            # and the HTML preview pipeline expose the absence of data clearly.
            path = output_base / "data_dictionary.xlsx"
            workbook = Workbook()
            summary = workbook.active
            summary.title = "Synthese"
            self._write_sheet(
                summary,
                self._data_dictionary_summary_headers(),
                [],
            )
            workbook.save(path)
            self.log(f"Data Dictionary genere (aucun objet detecte) : {path}")
            return [path]

        ordered_objects = sorted(
            objects, key=lambda obj: (obj.api_name or "").lower()
        )
        chunks: list[list[ObjectInfo]] = [
            ordered_objects[index : index + max_object_sheets]
            for index in range(0, len(ordered_objects), max_object_sheets)
        ]
        total_parts = len(chunks)
        written: list[Path] = []
        for part_index, chunk in enumerate(chunks, start=1):
            path = output_base / self._data_dictionary_filename(part_index)
            self._write_data_dictionary_workbook(
                chunk,
                path,
                part_index=part_index,
                total_parts=total_parts,
            )
            written.append(path)
        summary = (
            f"Data Dictionary genere ({len(ordered_objects)} objets, "
            f"{total_parts} fichier(s)) : "
            + ", ".join(path.name for path in written)
        )
        self.log(summary)
        return written

    @staticmethod
    def _data_dictionary_filename(part_index: int) -> str:
        if part_index <= 1:
            return "data_dictionary.xlsx"
        return f"data_dictionary_part_{part_index}.xlsx"

    @staticmethod
    def _data_dictionary_summary_headers() -> list[str]:
        return [
            "API Name",
            "Label",
            "Label pluriel",
            "Custom",
            "Modele de partage",
            "Statut deploiement",
            "Visibilite",
            "Nb champs",
            "Nb champs custom",
            "Nb record types",
            "Nb validation rules",
            "Nb relations",
            "Feuille",
            "Description",
        ]

    def _write_data_dictionary_workbook(
        self,
        objects_chunk: list[ObjectInfo],
        output_path: Path,
        *,
        part_index: int,
        total_parts: int,
    ) -> None:
        workbook = Workbook()
        used_names: set[str] = set()
        # Reserve the summary sheet name up front so no object collides with it.
        summary_name = self._unique_sheet_name("Synthese", used_names)
        summary = workbook.active
        summary.title = summary_name

        sheet_names_by_object: list[tuple[ObjectInfo, str]] = []
        for obj in objects_chunk:
            sheet_name = self._unique_sheet_name(
                obj.api_name or "Objet", used_names
            )
            sheet_names_by_object.append((obj, sheet_name))

        summary_rows = [
            [
                obj.api_name,
                obj.label,
                obj.plural_label,
                "Oui" if obj.custom else "Non",
                obj.sharing_model,
                obj.deployment_status,
                obj.visibility,
                len(obj.fields),
                sum(1 for field in obj.fields if field.custom),
                len(obj.record_types),
                len(obj.validation_rules),
                len(obj.relationships),
                sheet_name,
                obj.description,
            ]
            for obj, sheet_name in sheet_names_by_object
        ]
        self._write_sheet(
            summary,
            self._data_dictionary_summary_headers(),
            summary_rows,
        )

        if total_parts > 1:
            # Add a small indicator on the summary sheet so the user knows
            # other parts exist without having to open a file explorer.
            note_row = summary.max_row + 2
            summary.cell(
                row=note_row,
                column=1,
                value=(
                    f"Partie {part_index} / {total_parts}. "
                    "Les objets suivants se trouvent dans les autres fichiers "
                    "data_dictionary_part_*.xlsx."
                ),
            ).font = Font(italic=True)

        for obj, sheet_name in sheet_names_by_object:
            worksheet = workbook.create_sheet(sheet_name)
            self._write_object_fields_sheet(worksheet, obj)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _write_object_fields_sheet(self, worksheet, obj: ObjectInfo) -> None:
        headers = [
            "API Name",
            "Label",
            "Type",
            "Obligatoire",
            "Custom",
            "Reference vers",
            "Relationship Name",
            "Description",
        ]
        rows = [
            [
                field.api_name,
                field.label,
                field.data_type,
                "Oui" if field.required else "Non",
                "Oui" if field.custom else "Non",
                ", ".join(field.reference_to),
                field.relationship_name,
                field.description,
            ]
            for field in obj.fields
        ]
        self._write_sheet(worksheet, headers, rows)

        if not rows:
            # Leave a tiny hint explaining why the sheet is empty rather than
            # letting the user wonder if parsing failed.
            worksheet.cell(
                row=2,
                column=1,
                value="Aucun champ detecte dans la metadata pour cet objet.",
            ).font = Font(italic=True, color="666666")

    @staticmethod
    def _unique_sheet_name(desired: str, used: set[str]) -> str:
        """Return a unique, Excel-compliant sheet name and register it."""
        cleaned = _FORBIDDEN_SHEET_CHARS_RE.sub("_", desired or "").strip()
        cleaned = cleaned.strip("'")  # Excel rejects names wrapped in quotes
        if not cleaned:
            cleaned = "Feuille"
        base = cleaned[:31]
        candidate = base
        counter = 1
        # Case-insensitive comparison (Excel treats sheet names this way).
        existing_lower = {name.lower() for name in used}
        while candidate.lower() in existing_lower:
            counter += 1
            suffix = f"~{counter}"
            truncated = base[: max(1, 31 - len(suffix))]
            candidate = f"{truncated}{suffix}"
        used.add(candidate)
        return candidate

    def _write_sheet(self, worksheet, headers: list[str], rows: list[list[object]]) -> None:
        worksheet.append(headers)
        header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in rows:
            worksheet.append(row)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for index, header in enumerate(headers, start=1):
            max_length = len(header)
            for row in worksheet.iter_rows(min_col=index, max_col=index, min_row=2):
                value = row[0].value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 60)

    def _write_dict_sheet(self, worksheet, rows: list[dict[str, object]]) -> None:
        if rows:
            headers = list(rows[0].keys())
            data = [[row.get(header, "") for header in headers] for row in rows]
        else:
            headers = ["Information"]
            data = [["Aucune donnee trouvee"]]
        self._write_sheet(worksheet, headers, data)
