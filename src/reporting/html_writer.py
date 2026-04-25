from __future__ import annotations

import json
import os
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Callable

from src.analyzer.models import Finding
from src.core.models import (
    ApexArtifact,
    FlowInfo,
    MetadataSnapshot,
    ObjectInfo,
    PmdViolation,
    ReviewResult,
    SecurityArtifact,
    ValidationRuleInfo,
)
from src.core.utils import SF_NS, child_text, html_value, safe_slug, write_text
from src.reporting.formula_parser import FormulaNode, parse_formula
from src.reporting.html_mermaid import (
    data_transform_mermaid as _data_transform_mermaid_fn,
    data_transform_meta as _data_transform_meta_fn,
    mermaid_condition_label as _mermaid_condition_label_fn,
    mermaid_id as _mermaid_id_fn,
    mermaid_label as _mermaid_label_fn,
    object_mermaid as _object_mermaid_fn,
    short_error_text as _short_error_text_fn,
    validation_rule_mermaid as _validation_rule_mermaid_fn,
    wrap_mermaid_block as _wrap_mermaid_block_fn,
)


LogCallback = Callable[[str], None]


SEVERITY_CSS_CLASS: dict[str, str] = {
    "Critical": "sev-critical",
    "Major": "sev-major",
    "Minor": "sev-minor",
    "Info": "sev-info",
}

SEVERITY_LABEL: dict[str, str] = {
    "Critical": "Critique",
    "Major": "Majeur",
    "Minor": "Mineur",
    "Info": "Info",
}


MERMAID_RUNTIME_SCRIPT = r"""
<script src="https://cdn.jsdelivr.net/npm/mermaid@10.9.5/dist/mermaid.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/svg-pan-zoom@3.6.1/dist/svg-pan-zoom.min.js"></script>
<script>
(function(){
  function isVisible(el){
    var p = el.parentElement;
    while(p && p !== document.body){
      if(p.classList && p.classList.contains("tab-panel") && !p.classList.contains("active")){
        return false;
      }
      p = p.parentElement;
    }
    return true;
  }
  function captureSource(el){
    if(!el.hasAttribute("data-mermaid-source")){
      el.setAttribute("data-mermaid-source", el.textContent);
    }
  }
  function parseTranslate(str){
    if(!str) return {x:0,y:0};
    var m = str.match(/translate\(\s*([-\d.eE+]+)\s*[, ]\s*([-\d.eE+]+)\s*\)/);
    if(!m) return {x:0,y:0};
    return {x: parseFloat(m[1]), y: parseFloat(m[2])};
  }
  function centerOfNode(node){
    var tr = parseTranslate(node.getAttribute("transform"));
    return {x: tr.x, y: tr.y};
  }
  function shiftPathEndpoints(path, startDelta, endDelta){
    var d = path.getAttribute("d");
    if(!d) return;
    var numberRe = /-?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?/g;
    var nums = d.match(numberRe);
    if(!nums || nums.length < 4) return;
    if(startDelta){
      nums[0] = String(parseFloat(nums[0]) + startDelta.x);
      nums[1] = String(parseFloat(nums[1]) + startDelta.y);
    }
    if(endDelta){
      nums[nums.length-2] = String(parseFloat(nums[nums.length-2]) + endDelta.x);
      nums[nums.length-1] = String(parseFloat(nums[nums.length-1]) + endDelta.y);
    }
    var i = 0;
    var newD = d.replace(numberRe, function(){ return nums[i++]; });
    path.setAttribute("d", newD);
  }
  function nodeNameFromId(id){
    if(!id) return "";
    var m = id.match(/^flowchart-(.+?)-\d+$/);
    return m ? m[1] : id;
  }
  function findConnectedEdges(svg, nodeName){
    if(!nodeName) return {outgoing: [], incoming: []};
    var outgoing = [];
    var incoming = [];
    var paths = svg.querySelectorAll("g.edgePaths path, path.flowchart-link");
    paths.forEach(function(p){
      var cls = (p.getAttribute("class") || "").split(/\s+/);
      var fromCls = null, toCls = null;
      cls.forEach(function(c){
        if(c.indexOf("LS-") === 0){ fromCls = c.substring(3); }
        else if(c.indexOf("LE-") === 0){ toCls = c.substring(3); }
      });
      if(fromCls === nodeName){ outgoing.push(p); }
      if(toCls === nodeName){ incoming.push(p); }
    });
    return {outgoing: outgoing, incoming: incoming};
  }
  function moveEdgeLabel(svg, nodeName, delta){
    // labels are usually in g.edgeLabels and have labels per edge; we skip to avoid breakage.
  }
  function enableNodeDrag(svg, panZoom){
    var nodes = svg.querySelectorAll("g.node");
    nodes.forEach(function(node){
      if(node.dataset.mmDragEnabled === "true") return;
      node.dataset.mmDragEnabled = "true";
      var nodeName = nodeNameFromId(node.getAttribute("id"));
      var dragging = false;
      var startClient = null;
      var startOffset = null;
      var edges = null;
      var panEnabledBefore = true;
      node.addEventListener("mousedown", function(ev){
        if(ev.button !== 0) return;
        dragging = true;
        startClient = {x: ev.clientX, y: ev.clientY};
        startOffset = parseTranslate(node.getAttribute("transform"));
        edges = findConnectedEdges(svg, nodeName);
        try { panEnabledBefore = panZoom.isPanEnabled(); panZoom.disablePan(); } catch(e){}
        ev.stopPropagation();
        ev.preventDefault();
      });
      var lastDelta = {x:0,y:0};
      function onMove(ev){
        if(!dragging) return;
        var zoom = 1;
        try { zoom = panZoom.getZoom() || 1; } catch(e){}
        var dx = (ev.clientX - startClient.x) / zoom;
        var dy = (ev.clientY - startClient.y) / zoom;
        node.setAttribute("transform", "translate(" + (startOffset.x + dx) + "," + (startOffset.y + dy) + ")");
        var frameDelta = {x: dx - lastDelta.x, y: dy - lastDelta.y};
        if(edges){
          edges.outgoing.forEach(function(p){ shiftPathEndpoints(p, frameDelta, null); });
          edges.incoming.forEach(function(p){ shiftPathEndpoints(p, null, frameDelta); });
        }
        lastDelta = {x: dx, y: dy};
      }
      function onUp(){
        if(!dragging) return;
        dragging = false;
        lastDelta = {x:0, y:0};
        if(panEnabledBefore){ try { panZoom.enablePan(); } catch(e){} }
      }
      document.addEventListener("mousemove", onMove);
      document.addEventListener("mouseup", onUp);
    });
  }
  function enhanceContainer(container){
    if(container.dataset.mmEnhanced === "true") return;
    var mermaidDiv = container.querySelector(".mermaid");
    if(!mermaidDiv) return;
    if(mermaidDiv.getAttribute("data-processed") !== "true") return;
    var svg = mermaidDiv.querySelector("svg");
    if(!svg) return;
    container.dataset.mmEnhanced = "true";
    svg.removeAttribute("height");
    svg.removeAttribute("width");
    svg.style.width = "100%";
    svg.style.height = "100%";
    svg.style.maxWidth = "100%";
    var panZoom = null;
    try {
      panZoom = svgPanZoom(svg, {
        zoomEnabled: true,
        controlIconsEnabled: false,
        fit: true,
        center: true,
        minZoom: 0.2,
        maxZoom: 10,
        zoomScaleSensitivity: 0.35,
        dblClickZoomEnabled: false,
        preventMouseEventsDefault: false
      });
    } catch(err){
      console.error("svg-pan-zoom init failed", err);
      return;
    }
    var toolbar = container.querySelector(".mermaid-toolbar");
    if(toolbar){
      toolbar.addEventListener("click", function(ev){
        var btn = ev.target.closest("[data-mermaid-action]");
        if(!btn) return;
        var action = btn.getAttribute("data-mermaid-action");
        if(action === "zoom-in"){ panZoom.zoomBy(1.25); }
        else if(action === "zoom-out"){ panZoom.zoomBy(0.8); }
        else if(action === "reset"){ panZoom.resetZoom(); panZoom.center(); panZoom.fit(); }
      });
    }
    svg.addEventListener("dblclick", function(ev){
      if(ev.shiftKey){ panZoom.zoomBy(0.7); }
      else { panZoom.zoomBy(1.4); }
      ev.preventDefault();
    });
    enableNodeDrag(svg, panZoom);
  }
  function enhanceAll(scope){
    var containers = (scope || document).querySelectorAll(".mermaid-container");
    containers.forEach(enhanceContainer);
  }
  window.__enhanceMermaid = enhanceAll;
  window.__renderMermaid = function(root){
    if(!window.mermaid){return;}
    var scope = root || document;
    var nodes = Array.prototype.slice.call(scope.querySelectorAll(".mermaid"));
    var targets = nodes.filter(function(n){
      captureSource(n);
      if(!isVisible(n)){return false;}
      return n.getAttribute("data-processed") !== "true";
    });
    if(!targets.length){ enhanceAll(scope); return; }
    try{
      var result = window.mermaid.run({nodes: targets});
      if(result && typeof result.then === "function"){
        result.then(function(){ enhanceAll(scope); })
              .catch(function(e){ console.error("mermaid run", e); });
      } else {
        setTimeout(function(){ enhanceAll(scope); }, 50);
      }
    }
    catch(e){ console.error("mermaid run", e); }
  };
  function boot(){
    if(!window.mermaid){return;}
    try{
      window.mermaid.initialize({startOnLoad:false,securityLevel:"loose",theme:"default",flowchart:{htmlLabels:true,curve:"basis"}});
    }catch(e){ console.error("mermaid init", e); }
    Array.prototype.forEach.call(document.querySelectorAll(".mermaid"), captureSource);
    window.__renderMermaid();
  }
  if(document.readyState==="loading"){ document.addEventListener("DOMContentLoaded", boot); }
  else{ boot(); }
})();
</script>
""".strip()


class HtmlReportWriter:
    """Write the static HTML documentation site.

    The class owns the shared assets (CSS, icons, search index) and exposes
    ``write_*`` methods that emit one HTML page per object/class/trigger/
    flow. The orchestrator wires the writer up only when HTML output is
    enabled.
    """

    def __init__(
        self, output_dir: str | Path, log_callback: LogCallback | None = None
    ) -> None:
        self.output_dir = Path(output_dir)
        self.log: LogCallback = log_callback or (lambda message: None)
        self.assets_dir = self.output_dir / "assets"
        self.objects_dir = self.output_dir / "objects"
        self.apex_dir = self.output_dir / "apex"
        self.flows_dir = self.output_dir / "flows"
        self.omni_dir = self.output_dir / "omni"

    def write_assets(self) -> None:
        style = """
body { font-family: Arial, sans-serif; margin: 0; color: #1f2937; background: #f8fafc; }
.page { max-width: 1400px; margin: 0 auto; padding: 24px; }
.topnav { margin-bottom: 20px; }
.topnav a { text-decoration: none; color: #1d4ed8; }
h1, h2, h3 { color: #0f172a; }
.cards { display: flex; flex-wrap: wrap; gap: 16px; margin: 16px 0 24px; }
.card { background: white; border: 1px solid #cbd5e1; border-radius: 8px; padding: 16px; min-width: 180px; }
.card .value { display: block; font-size: 1.6rem; font-weight: bold; margin-top: 8px; }
table { width: 100%; border-collapse: collapse; background: white; margin: 12px 0 24px; }
th, td { border: 1px solid #cbd5e1; padding: 8px; text-align: left; vertical-align: top; }
th { background: #dbeafe; }
tr:nth-child(even) td { background: #f8fbff; }
.badge { display: inline-block; padding: 4px 10px; border-radius: 999px; background: #dbeafe; color: #1e3a8a; }
.badge.complexity-simple { background: #dcfce7; color: #166534; }
.badge.complexity-medium { background: #fef3c7; color: #92400e; }
.badge.complexity-complex { background: #fed7aa; color: #9a3412; }
.badge.complexity-very-complex { background: #fecaca; color: #991b1b; }
.section { margin-bottom: 28px; }
ul { background: white; border: 1px solid #cbd5e1; border-radius: 8px; padding: 16px 32px; }
.empty { color: #64748b; font-style: italic; }
.mermaid { background: white; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px; overflow: auto; }
.mermaid-container { margin: 14px 0; border: 1px solid #cbd5e1; border-radius: 10px; background: white; overflow: hidden; box-shadow: 0 1px 2px rgba(15,23,42,0.06); width: 100%; box-sizing: border-box; }
.mermaid-toolbar { display: flex; align-items: center; gap: 6px; padding: 6px 10px; border-bottom: 1px solid #e2e8f0; background: #f1f5f9; }
.mermaid-toolbar button.mm-btn { border: 1px solid #94a3b8; background: white; border-radius: 6px; padding: 3px 10px; cursor: pointer; font-weight: 600; color: #1e293b; min-width: 34px; line-height: 1.2; }
.mermaid-toolbar button.mm-btn:hover { background: #dbeafe; border-color: #60a5fa; }
.mermaid-toolbar .mm-hint { margin-left: auto; font-size: 0.78rem; color: #64748b; font-style: italic; }
.mermaid-container .mermaid { border: none; border-radius: 0; padding: 0; margin: 0; width: 100%; height: 720px; max-height: 85vh; overflow: hidden; background: white; user-select: none; box-sizing: border-box; display: block; position: relative; }
.mermaid-container .mermaid svg { width: 100% !important; height: 100% !important; max-width: none !important; display: block; }
.mermaid-container .mermaid g.node { cursor: grab; }
.mermaid-container .mermaid g.node:active { cursor: grabbing; }
.tab-panel .mermaid-container { max-width: 100%; }
code { background: #e2e8f0; padding: 2px 4px; border-radius: 4px; }
.smallcards .card { min-width: 150px; }
.graph-toolbar { display: flex; gap: 8px; margin-bottom: 10px; }
.graph-toolbar button { border: 1px solid #cbd5e1; background: #f8fafc; border-radius: 6px; padding: 6px 10px; cursor: pointer; }
.graph-toolbar button:hover { background: #e2e8f0; }
.graph-filters { display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 10px; }
.graph-filters label { display: inline-flex; align-items: center; gap: 6px; font-size: 0.92rem; color: #334155; }
.dependency-graph { height: 440px; border: 1px solid #cbd5e1; border-radius: 8px; background: #ffffff; }
.graph-legend { display: flex; flex-wrap: wrap; gap: 12px; margin: 8px 0 12px; }
.graph-legend .item { display: inline-flex; align-items: center; gap: 6px; font-size: 0.9rem; }
.findings-summary { display: flex; flex-wrap: wrap; gap: 10px; margin: 0 0 14px; }
.findings-summary .chip { display: inline-flex; align-items: center; gap: 6px; padding: 4px 12px; border-radius: 999px; font-weight: 600; font-size: 0.85rem; border: 1px solid #cbd5e1; background: white; color: #1e293b; }
.findings-summary .chip strong { font-weight: 700; }
.findings-summary .chip.sev-critical { background: #fef2f2; border-color: #f87171; color: #991b1b; }
.findings-summary .chip.sev-major { background: #fff7ed; border-color: #fb923c; color: #9a3412; }
.findings-summary .chip.sev-minor { background: #fefce8; border-color: #eab308; color: #854d0e; }
.findings-summary .chip.sev-info { background: #eff6ff; border-color: #60a5fa; color: #1e3a8a; }
.findings-list { list-style: none; padding: 0; margin: 0; background: white; border: 1px solid #cbd5e1; border-radius: 10px; overflow: hidden; }
.findings-list li.finding { padding: 14px 18px; border-top: 1px solid #e2e8f0; }
.findings-list li.finding:first-child { border-top: none; }
.findings-list li.finding .head { display: flex; flex-wrap: wrap; align-items: center; gap: 10px; margin-bottom: 6px; }
.findings-list li.finding .title { font-weight: 700; color: #0f172a; }
.findings-list li.finding .rule-id { font-family: ui-monospace, SFMono-Regular, Menlo, monospace; font-size: 0.85rem; color: #475569; }
.findings-list li.finding .sev-badge { display: inline-block; font-size: 0.75rem; padding: 2px 8px; border-radius: 999px; font-weight: 700; letter-spacing: 0.02em; }
.findings-list li.finding .sev-badge.sev-critical { background: #fee2e2; color: #991b1b; }
.findings-list li.finding .sev-badge.sev-major { background: #ffedd5; color: #9a3412; }
.findings-list li.finding .sev-badge.sev-minor { background: #fef9c3; color: #854d0e; }
.findings-list li.finding .sev-badge.sev-info { background: #dbeafe; color: #1e3a8a; }
.findings-list li.finding .category-badge { display: inline-block; font-size: 0.72rem; padding: 2px 8px; border-radius: 999px; background: #e2e8f0; color: #334155; letter-spacing: 0.02em; }
.findings-list li.finding .message { margin: 2px 0 6px; color: #1e293b; }
.findings-list li.finding .metadata { font-size: 0.85rem; color: #475569; margin: 4px 0; }
.findings-list li.finding .metadata dt { font-weight: 600; color: #0f172a; float: left; margin-right: 6px; }
.findings-list li.finding .metadata dd { margin: 0 0 4px; }
.findings-list li.finding .metadata a { color: #1d4ed8; }
.findings-list li.finding ul.details { margin: 6px 0 0; padding-left: 18px; }
.findings-list li.finding ul.details li { list-style: disc; color: #334155; }
.analyzer-summary-card { background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%); border: 1px solid #cbd5e1; border-radius: 12px; padding: 18px 22px; margin: 12px 0; }
.analyzer-summary-card h3 { margin-top: 0; }
.analyzer-summary-card table { width: 100%; border: none; background: transparent; border-collapse: collapse; }
.analyzer-summary-card table th, .analyzer-summary-card table td { border: none; background: transparent; padding: 6px 10px; }
.analyzer-summary-card table tbody tr:nth-child(even) td { background: #f1f5f9; }
.graph-legend .dot { width: 12px; height: 12px; border-radius: 999px; border: 1px solid #64748b; display: inline-block; }
.tabs { background: white; border: 1px solid #cbd5e1; border-radius: 8px; overflow: hidden; }
.tab-buttons { display: flex; flex-wrap: wrap; gap: 6px; padding: 10px; border-bottom: 1px solid #cbd5e1; background: #f8fafc; }
.tab-button { border: 1px solid #cbd5e1; border-radius: 999px; background: white; color: #334155; padding: 6px 12px; cursor: pointer; }
.tab-button.active { background: #dbeafe; color: #1e3a8a; border-color: #93c5fd; }
.tab-panel { display: none; padding: 14px; }
.tab-panel.active { display: block; }
        """.strip()
        write_text(self.assets_dir / "style.css", style)

    def write_object_pages(
        self,
        snapshot: MetadataSnapshot,
        *,
        analyzer_report=None,
    ) -> dict[str, Path]:
        output: dict[str, Path] = {}
        object_findings = getattr(analyzer_report, "objects", {}) if analyzer_report else {}
        validation_findings = getattr(analyzer_report, "validation_rules", {}) if analyzer_report else {}
        for item in snapshot.objects:
            path = self.objects_dir / f"{item.api_name}.html"
            vr_findings_for_object: list[Finding] = []
            for vr in item.validation_rules:
                key = f"{item.api_name}.{vr.full_name}"
                vr_findings_for_object.extend(validation_findings.get(key, []))
            content = self._render_object_page(
                item,
                snapshot,
                path,
                object_findings.get(item.api_name, []),
                vr_findings_for_object,
            )
            write_text(path, content)
            output[item.api_name] = path
        self.log(f"{len(output)} page(s) objet generee(s).")
        return output

    def write_apex_pages(
        self,
        snapshot: MetadataSnapshot,
        reviews: dict[str, ReviewResult],
        pmd_results: dict[str, list[PmdViolation]],
        *,
        analyzer_report=None,
    ) -> dict[str, Path]:
        artifacts = snapshot.apex_artifacts
        output: dict[str, Path] = {}
        for artifact in artifacts:
            filename = f"{safe_slug(artifact.name)}.html"
            output[artifact.name] = self.apex_dir / filename

        reference_index = self._build_apex_reference_index(artifacts)
        trigger_objects = {artifact.name: self._trigger_object_name(artifact) for artifact in artifacts}
        object_names = [item.api_name for item in snapshot.objects]
        flow_names = [item.name for item in snapshot.flows]
        apex_findings = getattr(analyzer_report, "apex", {}) if analyzer_report else {}
        for artifact in artifacts:
            path = output[artifact.name]
            dependencies = self._apex_dependencies(
                artifact,
                artifacts,
                reference_index,
                trigger_objects,
                object_names,
                flow_names,
            )
            write_text(
                path,
                self._render_apex_page(
                    artifact,
                    reviews[artifact.name],
                    path,
                    output,
                    dependencies,
                    pmd_results.get(artifact.name, []),
                    apex_findings.get(artifact.name, []),
                ),
            )
        self.log(f"{len(output)} page(s) Apex/Trigger generee(s).")
        return output

    def write_flow_pages(
        self,
        snapshot: MetadataSnapshot,
        reviews: dict[str, ReviewResult],
        object_pages: dict[str, Path],
        apex_pages: dict[str, Path],
        *,
        analyzer_report=None,
    ) -> dict[str, Path]:
        flows = snapshot.flows
        output: dict[str, Path] = {}
        for flow in flows:
            path = self.flows_dir / f"{safe_slug(flow.name)}.html"
            output[flow.name] = path

        flow_bodies = {
            flow.name: flow.source_path.read_text(encoding="utf-8", errors="ignore")
            if flow.source_path and flow.source_path.exists()
            else ""
            for flow in flows
        }
        flow_ref_index = self._build_flow_reference_index(flows, flow_bodies)
        object_names = [item.api_name for item in snapshot.objects]
        apex_names = [item.name for item in snapshot.apex_artifacts]
        flow_findings = getattr(analyzer_report, "flows", {}) if analyzer_report else {}

        for flow in flows:
            path = output[flow.name]
            dependencies = self._flow_dependencies(
                flow,
                flow_ref_index,
                flow_bodies.get(flow.name, ""),
                object_names,
                apex_names,
            )
            write_text(
                path,
                self._render_flow_page(
                    flow,
                    reviews[flow.name],
                    path,
                    dependencies,
                    output,
                    object_pages,
                    apex_pages,
                    flow_findings.get(flow.name, []),
                ),
            )
        self.log(f"{len(output)} page(s) Flow generee(s).")
        return output

    OMNI_SCORING_FOLDERS: dict[str, str] = {
        "omniscripts": "OmniScripts",
        "omniintegrationprocedures": "Integration Procedures",
        "omniuicard": "Omni UI Cards",
        "omnidatatransforms": "Data Transforms",
    }

    def write_omni_pages(
        self,
        snapshot: MetadataSnapshot,
        *,
        analyzer_report=None,
    ) -> dict[str, list[dict[str, object]]]:
        rows = snapshot.inventory.get("omnistudio") or []
        dt_findings = getattr(analyzer_report, "data_transforms", {}) if analyzer_report else {}
        grouped: dict[str, list[dict[str, object]]] = {}
        for row in rows:
            subcategory_raw = str(row.get("Dossier") or "").strip()
            label = self.OMNI_SCORING_FOLDERS.get(subcategory_raw.lower())
            if not label:
                continue
            grouped.setdefault(label, []).append(row)

        output: dict[str, list[dict[str, object]]] = {}
        total = 0
        for subcategory in sorted(grouped.keys(), key=lambda value: value.lower()):
            sub_slug = safe_slug(subcategory) or "autres"
            entries: list[dict[str, object]] = []
            used_slugs: set[str] = set()
            for row in sorted(grouped[subcategory], key=lambda item: str(item.get("Nom") or "").lower()):
                name = str(row.get("Nom") or "").strip() or "Sans nom"
                base_slug = safe_slug(name) or "composant"
                candidate_slug = base_slug
                counter = 2
                while candidate_slug in used_slugs:
                    candidate_slug = f"{base_slug}-{counter}"
                    counter += 1
                used_slugs.add(candidate_slug)

                page_path = self.omni_dir / sub_slug / f"{candidate_slug}.html"
                source_rel = str(row.get("Source") or "")
                content = self._render_omni_page(
                    name=name,
                    subcategory=subcategory,
                    row=row,
                    snapshot=snapshot,
                    current_path=page_path,
                    findings=dt_findings.get(name, []),
                )
                write_text(page_path, content)
                entries.append(
                    {
                        "name": name,
                        "page": page_path,
                        "source": source_rel,
                        "type": str(row.get("TypeFichier") or ""),
                    }
                )
                total += 1
            output[subcategory] = entries

        self.log(f"{total} page(s) OmniStudio generee(s).")
        return output

    def write_index(
        self,
        snapshot: MetadataSnapshot,
        object_pages: dict[str, Path],
        apex_pages: dict[str, Path],
        flow_pages: dict[str, Path],
        apex_reviews: dict[str, ReviewResult],
        flow_reviews: dict[str, ReviewResult],
        pmd_results: dict[str, list[PmdViolation]],
        omni_pages: dict[str, list[dict[str, object]]] | None = None,
        *,
        analyzer_report=None,
    ) -> Path:
        path = self.output_dir / "index.html"
        write_text(
            path,
            self._render_index(
                snapshot,
                object_pages,
                apex_pages,
                flow_pages,
                apex_reviews,
                flow_reviews,
                pmd_results,
                path,
                omni_pages or {},
                analyzer_report,
            ),
        )
        self.log(f"Index genere: {path}")
        return path

    def _render_object_page(
        self,
        item: ObjectInfo,
        snapshot: MetadataSnapshot,
        current_path: Path,
        object_findings: list[Finding] | None = None,
        validation_findings: list[Finding] | None = None,
    ) -> str:
        object_findings = object_findings or []
        validation_findings = validation_findings or []
        profiles = self._security_rows(snapshot.profiles, item.api_name)
        permsets = self._security_rows(snapshot.permission_sets, item.api_name)
        fields_rows = "".join(
            f"<tr><td>{html_value(field.api_name)}</td><td>{html_value(field.label)}</td>"
            f"<td>{html_value(field.data_type)}</td><td>{html_value(field.description)}</td>"
            f"<td>{'Oui' if field.required else 'Non'}</td></tr>"
            for field in item.fields
        ) or "<tr><td colspan='5' class='empty'>Aucun champ detecte.</td></tr>"

        record_type_rows = "".join(
            f"<tr><td>{html_value(record_type.full_name)}</td><td>{html_value(record_type.label)}</td>"
            f"<td>{html_value(record_type.description)}</td><td>{'Oui' if record_type.active else 'Non'}</td></tr>"
            for record_type in item.record_types
        ) or "<tr><td colspan='4' class='empty'>Aucun record type detecte.</td></tr>"

        description_rows = [
            ("Nom API", item.api_name),
            ("Label", item.label),
            ("Label pluriel", item.plural_label),
            ("Description", item.description),
            ("Deployment status", item.deployment_status),
            ("Sharing model", item.sharing_model),
            ("Visibilite", item.visibility),
        ]
        description_html = "".join(
            f"<li><strong>{html_value(label)}:</strong> {html_value(value or 'Non renseigne')}</li>"
            for label, value in description_rows
        )

        mermaid = self._object_mermaid(item)
        validation_rows = "".join(
            f"<tr><td>{html_value(vr.full_name)}</td><td>{'Oui' if vr.active else 'Non'}</td>"
            f"<td>{html_value(vr.description)}</td><td>{html_value(vr.error_display_field)}</td>"
            f"<td>{html_value(vr.error_message)}</td></tr>"
            for vr in item.validation_rules
        ) or "<tr><td colspan='5' class='empty'>Aucune regle de validation detectee.</td></tr>"

        validation_panels = []
        for vr in item.validation_rules:
            formula_html = f"<pre style='background:#f1f5f9; padding:12px; border-radius:6px; overflow:auto;'>{html_value(vr.error_condition_formula)}</pre>"
            mermaid_tree = self._validation_rule_mermaid(vr)
            validation_panels.append(
                f"<div class='section'><h3>{html_value(vr.full_name)}</h3>"
                f"<p><strong>Description:</strong> {html_value(vr.description or 'Non renseignee')}</p>"
                f"<p><strong>Message d'erreur:</strong> {html_value(vr.error_message)}</p>"
                f"<h4>Arbre de decision (Mermaid)</h4>{mermaid_tree}"
                f"<h4>Formule</h4>{formula_html}</div>"
            )
        validation_content = "".join(validation_panels) or "<p class='empty'>Aucune regle de validation detaillee.</p>"

        relation_table = "".join(
            f"<tr><td>{html_value(rel.field_name)}</td><td>{html_value(rel.relationship_type)}</td>"
            f"<td>{html_value(', '.join(rel.targets))}</td></tr>"
            for rel in item.relationships
        ) or "<tr><td colspan='3' class='empty'>Aucune relation detectee.</td></tr>"

        profile_rows = self._render_security_rows(profiles, "Aucun profil avec acces detecte.")
        permset_rows = self._render_security_rows(permsets, "Aucun permission set avec acces detecte.")

        combined_findings = list(object_findings) + list(validation_findings)
        analyzer_summary_inline = self._render_findings_summary(combined_findings)
        analyzer_content = self._render_analyzer_tab(combined_findings)

        synthesis_html = (
            "<ul>" + description_html + "</ul>"
            + "<div class='section'><h3>Alertes analyseur</h3>"
            + analyzer_summary_inline
            + "</div>"
        )

        tabs = self._tabbed_sections(
            f"object-{safe_slug(item.api_name)}",
            [
                ("Synthese", synthesis_html),
                ("Fields", f"<table><thead><tr><th>Name</th><th>Label</th><th>Type</th><th>Description</th><th>Required</th></tr></thead><tbody>{fields_rows}</tbody></table>"),
                ("Profiles", f"<table><thead><tr><th>Profile</th><th>Lecture</th><th>Creation</th><th>Modification</th><th>Suppression</th><th>Nb champs visibles</th><th>Nb champs modifiables</th></tr></thead><tbody>{profile_rows}</tbody></table>"),
                ("Permission Sets", f"<table><thead><tr><th>Permission Set</th><th>Lecture</th><th>Creation</th><th>Modification</th><th>Suppression</th><th>Nb champs visibles</th><th>Nb champs modifiables</th></tr></thead><tbody>{permset_rows}</tbody></table>"),
                ("Record Types", f"<table><thead><tr><th>Nom</th><th>Label</th><th>Description</th><th>Actif</th></tr></thead><tbody>{record_type_rows}</tbody></table>"),
                ("Validation Rules", f"<table><thead><tr><th>Nom</th><th>Actif</th><th>Description</th><th>Champ d'erreur</th><th>Message d'erreur</th></tr></thead><tbody>{validation_rows}</tbody></table><hr/>{validation_content}"),
                ("Relations", f"{mermaid}<table><thead><tr><th>Champ</th><th>Type</th><th>Cible</th></tr></thead><tbody>{relation_table}</tbody></table>"),
                ("Analyseur", analyzer_content),
            ],
        )
        body = f"""
{self._index_back_link(current_path, "objets")}
<h1>{html_value(item.api_name)}</h1>
<div class="cards">
  <div class="card"><span>Champs</span><span class="value">{len(item.fields)}</span></div>
  <div class="card"><span>Record types</span><span class="value">{len(item.record_types)}</span></div>
  <div class="card"><span>Regles de validation</span><span class="value">{len(item.validation_rules)}</span></div>
  <div class="card"><span>Relations</span><span class="value">{len(item.relationships)}</span></div>
</div>
{tabs}
"""
        return self._page(item.api_name, body, current_path)

    def _render_apex_page(
        self,
        artifact: ApexArtifact,
        review: ReviewResult,
        current_path: Path,
        apex_pages: dict[str, Path],
        dependencies: list[dict[str, str]],
        pmd_violations: list[PmdViolation],
        findings: list[Finding] | None = None,
    ) -> str:
        findings = findings or []
        metrics = "".join(
            f"<li><strong>{html_value(label)}:</strong> {html_value(value)}</li>"
            for label, value in review.metrics
        )
        improvements_for_heuristics = list(review.improvements) + self._findings_to_review_improvements(findings)
        positives = self._list_or_empty(review.positives, "Aucun point fort automatique detecte.")
        improvements = self._list_or_empty(improvements_for_heuristics, "Aucun point d'amelioration automatique detecte.")
        analyzer_tab = self._render_analyzer_tab(findings)
        analyzer_inline_summary = self._render_findings_summary(findings)
        code_preview = "\n".join(artifact.body.splitlines()[:120])
        dependency_rows = self._render_apex_dependency_rows(dependencies, current_path, apex_pages)
        dependency_graph = self._render_apex_dependency_graph(artifact, dependencies)
        pmd_rows = self._render_pmd_rows(pmd_violations)
        summary_html = (
            f"<p>{html_value(review.summary)}</p>"
            "<div class='section'><h3>Alertes analyseur</h3>"
            + analyzer_inline_summary
            + "</div>"
        )
        tabs = self._tabbed_sections(
            f"apex-{safe_slug(artifact.name)}",
            [
                ("Resume", summary_html),
                ("Metriques", f"<ul>{metrics}</ul>"),
                ("Points forts", positives),
                ("Heuristiques", improvements),
                ("Analyseur", analyzer_tab),
                (
                    "PMD",
                    f"<table><thead><tr><th>Regle</th><th>Ruleset</th><th>Priorite</th><th>Ligne</th><th>Message</th></tr></thead><tbody>{pmd_rows}</tbody></table>",
                ),
                (
                    "Liens",
                    f"<table><thead><tr><th>Composant lie</th><th>Categorie</th><th>Sous-type</th><th>Sens</th><th>Nature du lien</th></tr></thead><tbody>{dependency_rows}</tbody></table>",
                ),
                ("Graphe", dependency_graph),
                ("Extrait", f"<pre>{html_value(code_preview)}</pre>"),
            ],
        )
        body = f"""
{self._index_back_link(current_path, "apex-trigger")}
<h1>{html_value(artifact.name)}</h1>
<span class="badge">{html_value(artifact.kind)}</span>
{tabs}
"""
        return self._page(artifact.name, body, current_path)

    def _render_flow_page(
        self,
        flow: FlowInfo,
        review: ReviewResult,
        current_path: Path,
        dependencies: list[dict[str, str]],
        flow_pages: dict[str, Path],
        object_pages: dict[str, Path],
        apex_pages: dict[str, Path],
        findings: list[Finding] | None = None,
    ) -> str:
        findings = findings or []
        metrics = "".join(
            f"<li><strong>{html_value(label)}:</strong> {html_value(value)}</li>"
            for label, value in review.metrics
        )
        elements_rows = "".join(
            f"<tr><td>{html_value(element.element_type)}</td><td>{html_value(element.name)}</td>"
            f"<td>{html_value(element.label)}</td><td>{html_value(element.description)}</td><td>{html_value(element.target)}</td></tr>"
            for element in flow.elements
        ) or "<tr><td colspan='5' class='empty'>Aucun element detecte.</td></tr>"
        count_rows = "".join(
            f"<tr><td>{html_value(name)}</td><td>{count}</td></tr>"
            for name, count in sorted(flow.element_counts.items())
        ) or "<tr><td colspan='2' class='empty'>Aucun bloc detecte.</td></tr>"
        relation_rows = self._render_dependency_rows(
            dependencies,
            current_path,
            {"Flow": flow_pages, "Objet": object_pages, "Apex": apex_pages},
        )
        relation_graph = self._render_component_dependency_graph(flow.name, "Flow", dependencies, safe_slug(flow.name))
        analyzer_tab = self._render_analyzer_tab(findings)
        analyzer_inline_summary = self._render_findings_summary(findings)
        improvements_augmented = list(review.improvements) + self._findings_to_review_improvements(findings)
        summary_html = (
            f"<p>{html_value(review.summary)}</p>"
            "<div class='section'><h3>Alertes analyseur</h3>"
            + analyzer_inline_summary
            + "</div>"
        )
        tabs = self._tabbed_sections(
            f"flow-{safe_slug(flow.name)}",
            [
                ("Resume", summary_html),
                ("Metriques", f"<ul>{metrics}</ul>"),
                ("Repartition", f"<table><thead><tr><th>Type</th><th>Nombre</th></tr></thead><tbody>{count_rows}</tbody></table>"),
                ("Points forts", self._list_or_empty(review.positives, "Aucun point fort automatique detecte.")),
                ("Heuristiques", self._list_or_empty(improvements_augmented, "Aucun point d'amelioration automatique detecte.")),
                ("Analyseur", analyzer_tab),
                (
                    "Relations",
                    f"<table><thead><tr><th>Composant lie</th><th>Categorie</th><th>Sous-type</th><th>Sens</th><th>Nature du lien</th></tr></thead><tbody>{relation_rows}</tbody></table>{relation_graph}",
                ),
                ("Elements", f"<table><thead><tr><th>Type</th><th>Nom</th><th>Label</th><th>Description</th><th>Cible</th></tr></thead><tbody>{elements_rows}</tbody></table>"),
            ],
        )
        body = f"""
{self._index_back_link(current_path, "flows")}
<h1>{html_value(flow.name)}</h1>
<span class="badge">{html_value(flow.process_type or 'Flow')}</span>
<span class="badge {self._complexity_badge_class(flow.complexity_level)}">{html_value(flow.complexity_level)}</span>
<div class="cards smallcards">
  <div class="card"><span>Score complexite</span><span class="value">{flow.complexity_score}</span></div>
  <div class="card"><span>Elements</span><span class="value">{flow.total_elements}</span></div>
  <div class="card"><span>Documentes</span><span class="value">{flow.described_elements}</span></div>
  <div class="card"><span>Variables</span><span class="value">{flow.variable_total}</span></div>
  <div class="card"><span>Profondeur</span><span class="value">{flow.max_depth}</span></div>
  <div class="card"><span>Largeur max</span><span class="value">{flow.max_width}</span></div>
  <div class="card"><span>Hauteur min/max</span><span class="value">{flow.min_height}/{flow.max_height}</span></div>
</div>
{tabs}
"""
        return self._page(flow.name, body, current_path)

    def _render_omni_page(
        self,
        *,
        name: str,
        subcategory: str,
        row: dict[str, object],
        snapshot: MetadataSnapshot,
        current_path: Path,
        findings: list[Finding] | None = None,
    ) -> str:
        findings = findings or []
        source_rel = str(row.get("Source") or "")
        file_type = str(row.get("TypeFichier") or "")
        category_label = str(row.get("Categorie") or "OmniStudio")

        meta_rows = [
            ("Nom", name),
            ("Categorie", category_label),
            ("Sous-categorie", subcategory),
            ("Type de fichier", file_type),
            ("Source", source_rel or "Non renseigne"),
        ]
        meta_html = "".join(
            f"<li><strong>{html_value(label)}:</strong> {html_value(value or 'Non renseigne')}</li>"
            for label, value in meta_rows
        )

        preview_html = "<p class='empty'>Aucun contenu source exploitable.</p>"
        if source_rel:
            candidate = snapshot.source_dir / source_rel
            if candidate.exists() and candidate.is_file():
                try:
                    raw_text = candidate.read_text(encoding="utf-8", errors="ignore")
                except OSError:
                    raw_text = ""
                if raw_text:
                    lines = raw_text.splitlines()
                    preview_lines = lines[:200]
                    truncated = len(lines) > len(preview_lines)
                    preview_body = "\n".join(preview_lines)
                    suffix = "\n..." if truncated else ""
                    preview_html = f"<pre>{html_value(preview_body + suffix)}</pre>"

        folder = str(row.get("Dossier") or "").lower()
        is_data_transform = (
            "omnidatatransform" in folder
            or file_type.lower().endswith(".rpt-meta.xml")
            or subcategory.lower().replace("-", " ").strip() == "data transforms"
        )

        data_transform_payload: tuple[str, dict[str, str]] | None = None
        if is_data_transform and source_rel:
            candidate = snapshot.source_dir / source_rel
            if candidate.exists() and candidate.is_file():
                try:
                    xml_text = candidate.read_text(encoding="utf-8", errors="ignore")
                except OSError:
                    xml_text = ""
                if xml_text:
                    diagram_src = self._data_transform_mermaid(xml_text, name)
                    if diagram_src:
                        data_transform_payload = (
                            diagram_src,
                            self._data_transform_meta(xml_text),
                        )

        if data_transform_payload is not None:
            diagram_src, dt_meta = data_transform_payload
            wrapped = self._wrap_mermaid_block(diagram_src)
            mermaid_graph = (
                "<div class=\"section\">"
                f"<h3>Transformation de donnees : {html_value(name)}</h3>"
                "<p class='empty'>"
                f"Type : <strong>{html_value(dt_meta.get('type') or 'Non renseigne')}</strong> - "
                f"entrees <strong>{html_value(dt_meta.get('inputType') or '-')}</strong> "
                f"=&gt; sorties <strong>{html_value(dt_meta.get('outputType') or '-')}</strong>"
                "</p>"
                f"{wrapped}"
                "</div>"
            )
        else:
            msg_name = self._mermaid_label(name) or "Composant"
            msg_type = self._mermaid_label(file_type) or "Type"
            msg_sub = self._mermaid_label(subcategory) or "Sous-categorie"
            msg_cat = self._mermaid_label(category_label) or "Categorie"

            diagram_lines = [
                "flowchart LR",
                f'    Comp["{msg_name}"]',
                f'    Type["Type : {msg_type}"]',
                f'    Sub["Sous-categorie : {msg_sub}"]',
                f'    Cat["Categorie : {msg_cat}"]',
                "    Comp --> Type",
                "    Comp --> Sub",
                "    Comp --> Cat",
            ]
            omni_diagram = "\n".join(diagram_lines)
            wrapped = self._wrap_mermaid_block(omni_diagram)
            mermaid_graph = (
                "<div class=\"section\">"
                "<h3>Representation Graphique (OmniStudio)</h3>"
                f"{wrapped}"
                "</div>"
            )

        analyzer_tab = self._render_analyzer_tab(findings)
        analyzer_inline_summary = self._render_findings_summary(findings)
        synthesis_html = (
            f"<ul>{meta_html}</ul>"
            "<div class='section'><h3>Alertes analyseur</h3>"
            + analyzer_inline_summary
            + "</div>"
        )

        tabs = self._tabbed_sections(
            f"omni-{safe_slug(subcategory)}-{safe_slug(name)}",
            [
                ("Synthese", synthesis_html),
                ("Graphique", mermaid_graph),
                ("Analyseur", analyzer_tab),
                ("Contenu", preview_html),
            ],
        )
        body = f"""
{self._index_back_link(current_path, "omni")}
<h1>{html_value(name)}</h1>
<span class="badge">{html_value(category_label)}</span>
<span class="badge">{html_value(subcategory)}</span>
{tabs}
"""
        return self._page(name, body, current_path, include_mermaid=True)

    def _render_index(
        self,
        snapshot: MetadataSnapshot,
        object_pages: dict[str, Path],
        apex_pages: dict[str, Path],
        flow_pages: dict[str, Path],
        apex_reviews: dict[str, ReviewResult],
        flow_reviews: dict[str, ReviewResult],
        pmd_results: dict[str, list[PmdViolation]],
        current_path: Path,
        omni_pages: dict[str, list[dict[str, object]]],
        analyzer_report=None,
    ) -> str:
        metrics = snapshot.metrics
        object_rows = "".join(
            f"<tr><td><a href='{self._href(current_path, object_pages[item.api_name])}'>{html_value(item.api_name)}</a></td>"
            f"<td>{html_value(item.label)}</td><td>{len(item.fields)}</td><td>{len(item.relationships)}</td></tr>"
            for item in snapshot.objects
            if item.api_name in object_pages
        ) or "<tr><td colspan='4' class='empty'>Aucun objet analyse.</td></tr>"

        profile_rows = "".join(
            f"<tr><td>{html_value(item.name)}</td><td>{len(item.object_permissions)}</td><td>{len(item.field_permissions)}</td></tr>"
            for item in snapshot.profiles
        ) or "<tr><td colspan='3' class='empty'>Aucun profil analyse.</td></tr>"

        permset_rows = "".join(
            f"<tr><td>{html_value(item.name)}</td><td>{len(item.object_permissions)}</td><td>{len(item.field_permissions)}</td></tr>"
            for item in snapshot.permission_sets
        ) or "<tr><td colspan='3' class='empty'>Aucun permission set analyse.</td></tr>"

        apex_rows = "".join(
            f"<tr><td><a href='{self._href(current_path, apex_pages[item.name])}'>{html_value(item.name)}</a></td>"
            f"<td>{html_value(item.kind)}</td><td>{item.line_count}</td><td>{item.method_count}</td></tr>"
            for item in snapshot.apex_artifacts
            if item.name in apex_pages
        ) or "<tr><td colspan='4' class='empty'>Aucun artefact Apex analyse.</td></tr>"

        flow_rows = "".join(
            f"<tr><td><a href='{self._href(current_path, flow_pages[item.name])}'>{html_value(item.name)}</a></td>"
            f"<td>{html_value(item.process_type)}</td><td>{html_value(item.complexity_level)}</td><td>{item.complexity_score}</td><td>{item.total_elements}</td><td>{item.described_elements}</td></tr>"
            for item in snapshot.flows
            if item.name in flow_pages
        ) or "<tr><td colspan='6' class='empty'>Aucun flow analyse.</td></tr>"

        improvements_rows = self._render_index_improvements(
            snapshot,
            apex_reviews,
            flow_reviews,
            current_path,
            apex_pages,
            flow_pages,
        )
        pmd_rows = self._render_index_pmd_rows(
            snapshot,
            pmd_results,
            current_path,
            apex_pages,
        )
        excel_links = self._render_excel_exports(current_path)
        omni_panel = self._render_index_omni_panel(omni_pages, current_path)
        analyzer_panel = self._render_index_analyzer_panel(
            analyzer_report,
            current_path,
            object_pages,
            apex_pages,
            flow_pages,
        )

        tabs = self._tabbed_sections(
            "index",
            [
                (
                    "Exports Excel",
                    excel_links,
                ),
                (
                    "Omni",
                    omni_panel,
                ),
                (
                    "Objets",
                    f"<table><thead><tr><th>Objet</th><th>Label</th><th>Nb champs</th><th>Nb relations</th></tr></thead><tbody>{object_rows}</tbody></table>",
                ),
                (
                    "Profiles",
                    f"<table><thead><tr><th>Profile</th><th>Droits objet</th><th>Droits champ</th></tr></thead><tbody>{profile_rows}</tbody></table>",
                ),
                (
                    "Permission Sets",
                    f"<table><thead><tr><th>Permission Set</th><th>Droits objet</th><th>Droits champ</th></tr></thead><tbody>{permset_rows}</tbody></table>",
                ),
                (
                    "Apex / Trigger",
                    f"<table><thead><tr><th>Nom</th><th>Type</th><th>Lignes</th><th>Methodes</th></tr></thead><tbody>{apex_rows}</tbody></table>",
                ),
                (
                    "Flows",
                    f"<table><thead><tr><th>Nom</th><th>Type</th><th>Complexite</th><th>Score</th><th>Elements</th><th>Documentes</th></tr></thead><tbody>{flow_rows}</tbody></table>",
                ),
                (
                    "Analyseur",
                    analyzer_panel,
                ),
                (
                    "Ameliorations",
                    f"<table><thead><tr><th>Type</th><th>Composant</th><th>Amelioration</th></tr></thead><tbody>{improvements_rows}</tbody></table>",
                ),
                (
                    "Qualite PMD",
                    f"<table><thead><tr><th>Composant</th><th>Regle</th><th>Priorite</th><th>Ligne</th><th>Message</th></tr></thead><tbody>{pmd_rows}</tbody></table>",
                ),
            ],
        )
        omni_total = (
            metrics.omni_scripts
            + metrics.omni_integration_procedures
            + metrics.omni_ui_cards
            + metrics.omni_data_transforms
        )
        findings_card = ""
        findings_total = 0
        if analyzer_report is not None:
            findings_total = len(analyzer_report.all_findings())
            findings_card = (
                f'  <div class="card"><span>Findings analyseur</span>'
                f'<span class="value">{findings_total}</span></div>\n'
            )
        body = f"""
<h1>Documentation Salesforce</h1>
<p>Source analysee: <code>{html_value(snapshot.source_dir)}</code></p>
<div class="cards">
  <div class="card"><span>Niveau de customisation</span><span class="value">{html_value(metrics.level)}</span></div>
  <div class="card"><span>Score</span><span class="value">{metrics.score}</span></div>
  <div class="card"><span>Adopt vs Adapt</span><span class="value">{html_value(metrics.adopt_adapt_level)}</span></div>
  <div class="card"><span>Score Adopt vs Adapt</span><span class="value">{metrics.adopt_adapt_score}</span></div>
  <div class="card"><span>Objets custom</span><span class="value">{metrics.custom_objects}</span></div>
  <div class="card"><span>Champs custom</span><span class="value">{metrics.custom_fields}</span></div>
  <div class="card"><span>Flows</span><span class="value">{metrics.flows}</span></div>
  <div class="card"><span>Classes / Triggers</span><span class="value">{metrics.apex_classes + metrics.apex_triggers}</span></div>
  <div class="card"><span>Composants Omni</span><span class="value">{omni_total}</span></div>
{findings_card}</div>
{tabs}
"""
        return self._page("Index", body, current_path, include_mermaid=False)

    def _render_index_omni_panel(
        self,
        omni_pages: dict[str, list[dict[str, object]]],
        current_path: Path,
    ) -> str:
        if not omni_pages:
            return "<p class='empty'>Aucun composant OmniStudio detecte.</p>"

        sections: list[tuple[str, str]] = []
        for subcategory in sorted(omni_pages.keys(), key=lambda value: value.lower()):
            entries = omni_pages[subcategory]
            if not entries:
                rows = "<tr><td colspan='3' class='empty'>Aucun composant dans cette categorie.</td></tr>"
            else:
                rendered_rows: list[str] = []
                for entry in entries:
                    name = str(entry.get("name") or "")
                    page_path = entry.get("page")
                    source = str(entry.get("source") or "")
                    file_type = str(entry.get("type") or "")
                    if isinstance(page_path, Path):
                        link = f"<a href='{self._href(current_path, page_path)}'>{html_value(name)}</a>"
                    else:
                        link = html_value(name)
                    rendered_rows.append(
                        f"<tr><td>{link}</td><td>{html_value(file_type)}</td><td>{html_value(source)}</td></tr>"
                    )
                rows = "".join(rendered_rows)

            label = f"{subcategory} ({len(entries)})"
            table = (
                "<table><thead><tr><th>Composant</th><th>Type</th><th>Source</th></tr></thead>"
                f"<tbody>{rows}</tbody></table>"
            )
            sections.append((label, table))

        return self._tabbed_sections("index-omni", sections)

    def _render_index_analyzer_panel(
        self,
        analyzer_report,
        current_path: Path,
        object_pages: dict[str, Path],
        apex_pages: dict[str, Path],
        flow_pages: dict[str, Path],
    ) -> str:
        if analyzer_report is None:
            return "<p class='empty'>Analyseur non execute.</p>"

        findings = analyzer_report.all_findings()
        summary = self._render_findings_summary(findings)
        if not findings:
            return summary + "<p class='empty'>Aucun finding : le projet respecte toutes les regles activees.</p>"

        rule_counts = analyzer_report.rule_counts()
        rules_by_id = {rule.id: rule for rule in analyzer_report.rules_used}
        rule_rows = []
        for rule_id, count in sorted(rule_counts.items(), key=lambda kv: (-kv[1], kv[0])):
            rule = rules_by_id.get(rule_id)
            if not rule:
                continue
            sev_css = SEVERITY_CSS_CLASS.get(rule.severity, "sev-info")
            sev_label = SEVERITY_LABEL.get(rule.severity, rule.severity)
            reference = ""
            if rule.reference:
                reference = f"<a href='{html_value(rule.reference)}' target='_blank' rel='noopener'>Reference</a>"
            rule_rows.append(
                f"<tr>"
                f"<td><span class='sev-badge {sev_css}'>{html_value(sev_label)}</span></td>"
                f"<td>{html_value(rule.id)}</td>"
                f"<td>{html_value(rule.title)}</td>"
                f"<td>{html_value(rule.category)} - {html_value(rule.subcategory)}</td>"
                f"<td>{count}</td>"
                f"<td>{reference}</td>"
                f"</tr>"
            )
        rule_table = (
            "<table><thead><tr><th>Severite</th><th>Identifiant</th><th>Regle</th><th>Categorie</th><th>Occurrences</th><th>Reference</th></tr></thead>"
            f"<tbody>{''.join(rule_rows)}</tbody></table>"
        )

        # Tableau par artefact avec lien vers la page
        artifact_rows: list[str] = []

        def _artifact_row(kind: str, name: str, findings_list: list[Finding], href: str) -> str:
            if not findings_list:
                return ""
            counts = {"Critical": 0, "Major": 0, "Minor": 0, "Info": 0}
            for finding in findings_list:
                counts[finding.rule.severity] = counts.get(finding.rule.severity, 0) + 1
            sev_cells = "".join(
                f"<td>{counts.get(sev, 0)}</td>"
                for sev in ("Critical", "Major", "Minor", "Info")
            )
            name_cell = (
                f"<a href='{html_value(href)}'>{html_value(name)}</a>" if href else html_value(name)
            )
            return (
                f"<tr><td>{html_value(kind)}</td><td>{name_cell}</td>"
                f"{sev_cells}<td>{len(findings_list)}</td></tr>"
            )

        for name, flist in analyzer_report.objects.items():
            page = object_pages.get(name)
            href = self._href(current_path, page) if page else ""
            row = _artifact_row("Objet", name, flist, href)
            if row:
                artifact_rows.append(row)

        for name, flist in analyzer_report.apex.items():
            page = apex_pages.get(name)
            href = self._href(current_path, page) if page else ""
            row = _artifact_row("Apex", name, flist, href)
            if row:
                artifact_rows.append(row)

        for name, flist in analyzer_report.flows.items():
            page = flow_pages.get(name)
            href = self._href(current_path, page) if page else ""
            row = _artifact_row("Flow", name, flist, href)
            if row:
                artifact_rows.append(row)

        for name, flist in analyzer_report.validation_rules.items():
            row = _artifact_row("Validation Rule", name, flist, "")
            if row:
                artifact_rows.append(row)

        for name, flist in analyzer_report.data_transforms.items():
            row = _artifact_row("Data Transform", name, flist, "")
            if row:
                artifact_rows.append(row)

        artifact_rows.sort()
        if not artifact_rows:
            artifact_table = "<p class='empty'>Aucun composant impacte.</p>"
        else:
            artifact_table = (
                "<table><thead><tr><th>Type</th><th>Composant</th>"
                "<th>Critique</th><th>Majeur</th><th>Mineur</th><th>Info</th><th>Total</th></tr></thead>"
                f"<tbody>{''.join(artifact_rows)}</tbody></table>"
            )

        note = (
            "<p class='empty'>Analyseur inspire de "
            "<a href='https://docs.pmd-code.org/latest/pmd_rules_apex.html' target='_blank' rel='noopener'>PMD Apex</a>, du "
            "<a href='https://architect.salesforce.com/docs/architect/well-architected/guide/overview.html' target='_blank' rel='noopener'>Salesforce Well-Architected Framework</a>, "
            "des <a href='https://architect.salesforce.com/decision-guides' target='_blank' rel='noopener'>Decision Guides Salesforce</a> et des bonnes pratiques "
            "<a href='https://admin.salesforce.com/' target='_blank' rel='noopener'>Salesforce Admins</a>. "
            "Les regles sont declarees dans <code>src/analyzer/rules.xml</code> et peuvent etre activees / desactivees via l'attribut <code>enabled</code>.</p>"
        )

        sections = [
            ("Synthese par regle", rule_table),
            ("Par composant", artifact_table),
        ]
        return summary + self._tabbed_sections("index-analyzer", sections) + note

    def _render_excel_exports(self, current_path: Path) -> str:
        excel_dir = self.output_dir / "excel"
        if not excel_dir.exists():
            return "<p class='empty'>Aucun export Excel detecte.</p>"

        files = sorted(excel_dir.glob("*.xlsx"), key=lambda path: path.name.lower())
        if not files:
            return "<p class='empty'>Aucun export Excel detecte.</p>"

        rows: list[str] = []
        for file_path in files:
            xlsx_href = self._href(current_path, file_path)
            preview_path = file_path.with_suffix(".html")
            if preview_path.exists():
                preview_href = self._href(current_path, preview_path)
                preview_cell = (
                    f"<a href='{preview_href}'>{html_value(file_path.stem)}</a>"
                )
            else:
                preview_cell = (
                    f"<span class='empty'>{html_value(file_path.stem)}</span>"
                )
            rows.append(
                f"<tr><td>{preview_cell}</td>"
                f"<td><a href='{xlsx_href}'>{html_value(file_path.name)}</a></td></tr>"
            )
        return (
            "<table><thead><tr><th>Apercu HTML</th><th>Fichier Excel</th></tr></thead>"
            f"<tbody>{''.join(rows)}</tbody></table>"
        )

    def write_excel_preview_pages(self) -> dict[Path, Path]:
        excel_dir = self.output_dir / "excel"
        mapping: dict[Path, Path] = {}
        if not excel_dir.exists():
            return mapping

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.log(f"openpyxl indisponible, apercu Excel ignore: {exc}")
            return mapping

        for xlsx_path in sorted(excel_dir.glob("*.xlsx")):
            try:
                workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            except Exception as exc:
                self.log(f"Echec lecture {xlsx_path.name}: {exc}")
                continue
            try:
                sections = self._build_excel_preview_sections(workbook)
            finally:
                workbook.close()
            preview_path = xlsx_path.with_suffix(".html")
            tabs = self._tabbed_sections(
                f"excel-preview-{safe_slug(xlsx_path.stem)}", sections
            )
            xlsx_href = self._href(preview_path, xlsx_path)
            body = f"""
{self._index_back_link(preview_path, "exports-excel")}
<h1>{html_value(xlsx_path.name)}</h1>
<p>Apercu des feuilles du classeur. <a href="{xlsx_href}">Telecharger le fichier Excel original</a>.</p>
{tabs}
""".strip()
            write_text(
                preview_path,
                self._page(xlsx_path.stem, body, preview_path, include_mermaid=False),
            )
            mapping[xlsx_path] = preview_path
        if mapping:
            self.log(f"{len(mapping)} apercu(s) HTML de classeur Excel genere(s).")
        return mapping

    def _build_excel_preview_sections(self, workbook) -> list[tuple[str, str]]:
        max_rows = 500
        sections: list[tuple[str, str]] = []
        for sheet in workbook.worksheets:
            headers: tuple | None = None
            body_rows: list[str] = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    body_rows.append(
                        f"<tr><td colspan='{len(headers) if headers else 1}' "
                        "class='empty'>... (apercu tronque a "
                        f"{max_rows} lignes) ...</td></tr>"
                    )
                    break
                if index == 0:
                    headers = row
                    continue
                cells = "".join(
                    f"<td>{html_value('' if value is None else value)}</td>"
                    for value in row
                )
                body_rows.append(f"<tr>{cells}</tr>")

            if headers:
                header_cells = "".join(
                    f"<th>{html_value('' if value is None else value)}</th>"
                    for value in headers
                )
                table = (
                    f"<table><thead><tr>{header_cells}</tr></thead>"
                    f"<tbody>{''.join(body_rows)}</tbody></table>"
                )
            else:
                table = "<p class='empty'>Feuille vide.</p>"

            label = sheet.title or "Feuille"
            sections.append((label, table))

        if not sections:
            sections.append(
                ("Contenu", "<p class='empty'>Aucune feuille exploitable.</p>")
            )
        return sections

    def _render_index_improvements(
        self,
        snapshot: MetadataSnapshot,
        apex_reviews: dict[str, ReviewResult],
        flow_reviews: dict[str, ReviewResult],
        current_path: Path,
        apex_pages: dict[str, Path],
        flow_pages: dict[str, Path],
    ) -> str:
        rows: list[str] = []
        for artifact in snapshot.apex_artifacts:
            review = apex_reviews.get(artifact.name)
            if review is None:
                continue
            page = apex_pages.get(artifact.name)
            if page:
                component = f"<a href='{self._href(current_path, page)}'>{html_value(artifact.name)}</a>"
            else:
                component = html_value(artifact.name)
            for improvement in review.improvements:
                rows.append(
                    f"<tr><td>Apex/{html_value(artifact.kind)}</td><td>{component}</td><td>{html_value(improvement)}</td></tr>"
                )

        for flow in snapshot.flows:
            review = flow_reviews.get(flow.name)
            if review is None:
                continue
            page = flow_pages.get(flow.name)
            if page:
                component = f"<a href='{self._href(current_path, page)}'>{html_value(flow.name)}</a>"
            else:
                component = html_value(flow.name)
            for improvement in review.improvements:
                rows.append(
                    f"<tr><td>Flow</td><td>{component}</td><td>{html_value(improvement)}</td></tr>"
                )

        return "".join(rows) or "<tr><td colspan='3' class='empty'>Aucune amelioration detectee.</td></tr>"

    def _render_index_pmd_rows(
        self,
        snapshot: MetadataSnapshot,
        pmd_results: dict[str, list[PmdViolation]],
        current_path: Path,
        apex_pages: dict[str, Path],
    ) -> str:
        rows: list[str] = []
        for artifact in snapshot.apex_artifacts:
            violations = pmd_results.get(artifact.name, [])
            if not violations:
                continue
            target = apex_pages.get(artifact.name)
            component = (
                f"<a href='{self._href(current_path, target)}'>{html_value(artifact.name)}</a>"
                if target
                else html_value(artifact.name)
            )
            for violation in violations:
                line_value = violation.begin_line or ""
                rows.append(
                    f"<tr><td>{component}</td><td>{html_value(violation.rule)}</td>"
                    f"<td>{html_value(violation.priority)}</td><td>{html_value(line_value)}</td>"
                    f"<td>{html_value(violation.message)}</td></tr>"
                )
        return "".join(rows) or "<tr><td colspan='5' class='empty'>Aucune violation PMD detectee.</td></tr>"

    def _security_rows(self, artifacts: list[SecurityArtifact], object_name: str) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        prefix = f"{object_name}."
        for artifact in artifacts:
            permission = next(
                (item for item in artifact.object_permissions if item.object_name == object_name),
                None,
            )
            visible_fields = sum(1 for item in artifact.field_permissions if item.field_name.startswith(prefix) and item.readable)
            editable_fields = sum(1 for item in artifact.field_permissions if item.field_name.startswith(prefix) and item.editable)
            if permission or visible_fields or editable_fields:
                rows.append(
                    {
                        "name": artifact.name,
                        "read": "Oui" if permission and permission.allow_read else "Non",
                        "create": "Oui" if permission and permission.allow_create else "Non",
                        "edit": "Oui" if permission and permission.allow_edit else "Non",
                        "delete": "Oui" if permission and permission.allow_delete else "Non",
                        "visible_fields": visible_fields,
                        "editable_fields": editable_fields,
                    }
                )
        return rows

    def _render_security_rows(self, rows: list[dict[str, object]], empty_text: str) -> str:
        if not rows:
            return f"<tr><td colspan='7' class='empty'>{html_value(empty_text)}</td></tr>"
        return "".join(
            f"<tr><td>{html_value(row['name'])}</td><td>{html_value(row['read'])}</td>"
            f"<td>{html_value(row['create'])}</td><td>{html_value(row['edit'])}</td>"
            f"<td>{html_value(row['delete'])}</td><td>{html_value(row['visible_fields'])}</td>"
            f"<td>{html_value(row['editable_fields'])}</td></tr>"
            for row in rows
        )

    def _render_pmd_rows(self, violations: list[PmdViolation]) -> str:
        if not violations:
            return "<tr><td colspan='5' class='empty'>Aucune violation PMD detectee.</td></tr>"
        rows = []
        for violation in violations:
            line_display = (
                f"{violation.begin_line}-{violation.end_line}"
                if violation.begin_line and violation.end_line and violation.end_line != violation.begin_line
                else str(violation.begin_line or "")
            )
            rows.append(
                f"<tr><td>{html_value(violation.rule)}</td><td>{html_value(violation.ruleset)}</td>"
                f"<td>{html_value(violation.priority)}</td><td>{html_value(line_display)}</td>"
                f"<td>{html_value(violation.message)}</td></tr>"
            )
        return "".join(rows)

    def _render_findings_summary(self, findings: list[Finding]) -> str:
        if not findings:
            return "<p class='empty'>Aucun point d'alerte detecte par l'analyseur.</p>"
        counts: dict[str, int] = {}
        for finding in findings:
            counts[finding.rule.severity] = counts.get(finding.rule.severity, 0) + 1
        chips = []
        for severity in ("Critical", "Major", "Minor", "Info"):
            if counts.get(severity):
                css = SEVERITY_CSS_CLASS.get(severity, "")
                label = SEVERITY_LABEL.get(severity, severity)
                chips.append(
                    f"<span class='chip {css}'><strong>{counts[severity]}</strong> {html_value(label)}</span>"
                )
        return "<div class='findings-summary'>" + "".join(chips) + "</div>"

    def _render_findings_list(self, findings: list[Finding]) -> str:
        if not findings:
            return "<p class='empty'>Aucun point d'alerte detecte par l'analyseur.</p>"
        items: list[str] = []
        for finding in findings:
            rule = finding.rule
            severity_css = SEVERITY_CSS_CLASS.get(rule.severity, "sev-info")
            severity_label = SEVERITY_LABEL.get(rule.severity, rule.severity)
            reference = ""
            if rule.reference:
                reference = (
                    f"<dt>Reference:</dt><dd><a href='{html_value(rule.reference)}' target='_blank' rel='noopener'>{html_value(rule.reference)}</a></dd>"
                )
            details_html = ""
            if finding.details:
                detail_items = "".join(
                    f"<li>{html_value(detail)}</li>" for detail in finding.details
                )
                details_html = f"<ul class='details'>{detail_items}</ul>"
            subcat = f" - {html_value(rule.subcategory)}" if rule.subcategory else ""
            items.append(
                "<li class='finding'>"
                "<div class='head'>"
                f"<span class='sev-badge {severity_css}'>{html_value(severity_label)}</span>"
                f"<span class='category-badge'>{html_value(rule.category)}{subcat}</span>"
                f"<span class='rule-id'>{html_value(rule.id)}</span>"
                f"<span class='title'>{html_value(rule.title)}</span>"
                "</div>"
                f"<div class='message'>{html_value(finding.message or rule.description)}</div>"
                "<dl class='metadata'>"
                f"<dt>Justification:</dt><dd>{html_value(rule.rationale)}</dd>"
                f"<dt>Remediation:</dt><dd>{html_value(rule.remediation)}</dd>"
                f"<dt>Source:</dt><dd>{html_value(rule.source)}</dd>"
                f"{reference}"
                "</dl>"
                f"{details_html}"
                "</li>"
            )
        return "<ul class='findings-list'>" + "".join(items) + "</ul>"

    def _findings_to_review_improvements(self, findings: list[Finding]) -> list[str]:
        lines: list[str] = []
        for finding in findings:
            severity_label = SEVERITY_LABEL.get(finding.rule.severity, finding.rule.severity)
            lines.append(
                f"[{severity_label}] {finding.rule.id} - {finding.rule.title} : {finding.message or finding.rule.description}"
            )
        return lines

    def _render_analyzer_tab(self, findings: list[Finding]) -> str:
        summary = self._render_findings_summary(findings)
        body = self._render_findings_list(findings)
        note = (
            "<p class='empty'>Regles inspirees de PMD Apex, du Salesforce Well-Architected Framework et des guides "
            "Salesforce Architects / Admins. Chaque regle peut etre activee ou desactivee dans "
            "<code>src/analyzer/rules.xml</code>.</p>"
        )
        return summary + body + note

    # ------------------------------------------------------------------
    # Mermaid helpers (implementation lives in src.reporting.html_mermaid)
    # ------------------------------------------------------------------

    def _wrap_mermaid_block(self, diagram_source: str) -> str:
        return _wrap_mermaid_block_fn(diagram_source)

    def _data_transform_meta(self, xml_content: str) -> dict[str, str]:
        return _data_transform_meta_fn(xml_content)

    def _data_transform_mermaid(self, xml_content: str, name: str) -> str | None:
        return _data_transform_mermaid_fn(xml_content, name)

    def _mermaid_condition_label(self, text: str) -> str:
        return _mermaid_condition_label_fn(text)

    def _short_error_text(self, text: str, limit: int = 60) -> str:
        return _short_error_text_fn(text, limit)

    def _validation_rule_mermaid(self, vr: ValidationRuleInfo) -> str:
        return _validation_rule_mermaid_fn(vr)

    def _object_mermaid(self, item: ObjectInfo) -> str:
        return _object_mermaid_fn(item)

    def _mermaid_id(self, value: str) -> str:
        return _mermaid_id_fn(value)

    def _mermaid_label(self, value: str) -> str:
        return _mermaid_label_fn(value)

    def _list_or_empty(self, items: list[str], empty_text: str) -> str:
        if not items:
            return f"<p class='empty'>{html_value(empty_text)}</p>"
        return "<ul>" + "".join(f"<li>{html_value(item)}</li>" for item in items) + "</ul>"

    def _complexity_badge_class(self, level: str) -> str:
        mapping = {
            "Simple": "complexity-simple",
            "Moyen": "complexity-medium",
            "Complexe": "complexity-complex",
            "Tres complexe": "complexity-very-complex",
        }
        return mapping.get(level, "")

    def _build_apex_reference_index(self, artifacts: list[ApexArtifact]) -> dict[str, set[str]]:
        references: dict[str, set[str]] = {}
        patterns = {
            artifact.name: re.compile(rf"\b{re.escape(artifact.name)}\b", re.IGNORECASE)
            for artifact in artifacts
        }
        for source in artifacts:
            linked: set[str] = set()
            for target in artifacts:
                if target.name == source.name:
                    continue
                if patterns[target.name].search(source.body):
                    linked.add(target.name)
            references[source.name] = linked
        return references

    def _trigger_object_name(self, artifact: ApexArtifact) -> str:
        if artifact.kind != "trigger":
            return ""
        match = re.search(r"(?im)^\s*trigger\s+\w+\s+on\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", artifact.body)
        return match.group(1) if match else ""

    def _apex_dependencies(
        self,
        artifact: ApexArtifact,
        artifacts: list[ApexArtifact],
        reference_index: dict[str, set[str]],
        trigger_objects: dict[str, str],
        object_names: list[str],
        flow_names: list[str],
    ) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        by_name = {item.name: item for item in artifacts}
        seen: set[tuple[str, str, str, str]] = set()

        for target_name in sorted(reference_index.get(artifact.name, set()), key=str.lower):
            target = by_name.get(target_name)
            if target is None:
                continue
            key = (target_name, "Sortant", "Reference code", "Apex")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": target_name,
                    "category": "Apex",
                    "subtype": target.kind,
                    "direction": "Sortant",
                    "relation": "Reference code",
                }
            )

        for source_name, targets in reference_index.items():
            if source_name == artifact.name or artifact.name not in targets:
                continue
            source = by_name.get(source_name)
            if source is None:
                continue
            key = (source_name, "Entrant", "Reference code", "Apex")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": source_name,
                    "category": "Apex",
                    "subtype": source.kind,
                    "direction": "Entrant",
                    "relation": "Reference code",
                }
            )

        if artifact.kind == "trigger":
            current_object = trigger_objects.get(artifact.name, "")
            if current_object:
                for target in artifacts:
                    if target.name == artifact.name or target.kind != "trigger":
                        continue
                    if trigger_objects.get(target.name) != current_object:
                        continue
                    key = (target.name, "Sortant", f"Meme objet trigger ({current_object})", "Apex")
                    if key in seen:
                        continue
                    seen.add(key)
                    rows.append(
                        {
                            "name": target.name,
                            "category": "Apex",
                            "subtype": target.kind,
                            "direction": "Sortant",
                            "relation": f"Meme objet trigger ({current_object})",
                        }
                    )

        for object_name in sorted(object_names, key=str.lower):
            if not re.search(rf"\b{re.escape(object_name)}\b", artifact.body):
                continue
            key = (object_name, "Sortant", "Usage objet", "Objet")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": object_name,
                    "category": "Objet",
                    "subtype": "sObject",
                    "direction": "Sortant",
                    "relation": "Usage objet",
                }
            )

        for flow_name in sorted(flow_names, key=str.lower):
            if not re.search(rf"\b{re.escape(flow_name)}\b", artifact.body):
                continue
            key = (flow_name, "Sortant", "Reference flow", "Flow")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": flow_name,
                    "category": "Flow",
                    "subtype": "Flow",
                    "direction": "Sortant",
                    "relation": "Reference flow",
                }
            )

        metadata_matches = sorted(set(re.findall(r"\b([A-Za-z_][A-Za-z0-9_]*__mdt)\b", artifact.body)))
        for metadata_name in metadata_matches:
            key = (metadata_name, "Sortant", "Reference metadata", "Metadata")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": metadata_name,
                    "category": "Metadata",
                    "subtype": "CustomMetadata",
                    "direction": "Sortant",
                    "relation": "Reference metadata",
                }
            )

        rows.sort(key=lambda item: (item["category"], item["direction"], item["name"].lower()))
        return rows

    def _render_apex_dependency_rows(
        self,
        rows: list[dict[str, str]],
        current_path: Path,
        apex_pages: dict[str, Path],
    ) -> str:
        return self._render_dependency_rows(rows, current_path, {"Apex": apex_pages})

    def _render_apex_dependency_graph(
        self,
        artifact: ApexArtifact,
        rows: list[dict[str, str]],
    ) -> str:
        return self._render_component_dependency_graph(artifact.name, "Apex", rows, safe_slug(artifact.name))

    def _render_dependency_rows(
        self,
        rows: list[dict[str, str]],
        current_path: Path,
        link_maps: dict[str, dict[str, Path]],
    ) -> str:
        if not rows:
            return "<tr><td colspan='5' class='empty'>Aucun lien detecte.</td></tr>"

        rendered_rows: list[str] = []
        for row in rows:
            target_path = link_maps.get(row["category"], {}).get(row["name"])
            if target_path:
                name_html = (
                    f"<a href='{self._href(current_path, target_path)}'>{html_value(row['name'])}</a>"
                )
            else:
                name_html = html_value(row["name"])
            rendered_rows.append(
                f"<tr><td>{name_html}</td><td>{html_value(row['category'])}</td><td>{html_value(row['subtype'])}</td>"
                f"<td>{html_value(row['direction'])}</td><td>{html_value(row['relation'])}</td></tr>"
            )
        return "".join(rendered_rows)

    def _render_component_dependency_graph(
        self,
        center_name: str,
        center_category: str,
        rows: list[dict[str, str]],
        key_suffix: str,
    ) -> str:
        if not rows:
            return "<p class='empty'>Aucun graphe a afficher, aucun lien n'a ete detecte.</p>"

        nodes: dict[str, dict[str, object]] = {
            center_name: {
                "id": center_name,
                "label": center_name,
                "title": f"{center_category}: {center_name}",
                "color": {"background": "#bfdbfe", "border": "#2563eb"},
                "shape": "box",
                "componentKind": center_category,
                "category": center_category,
            }
        }
        edges: list[dict[str, str]] = []
        edge_seen: set[tuple[str, str, str]] = set()
        for row in rows:
            target_name = row["name"]
            if target_name not in nodes:
                nodes[target_name] = {
                    "id": target_name,
                    "label": target_name,
                    "title": f"{row['category']} - {row['subtype']}: {target_name}",
                    "shape": "box",
                    "componentKind": row["subtype"],
                    "category": row["category"],
                    "color": self._dependency_node_color(row["category"]),
                }
            if row["direction"] == "Entrant":
                source = target_name
                destination = center_name
            else:
                source = center_name
                destination = target_name
            edge_key = (source, destination, row["relation"])
            if edge_key in edge_seen:
                continue
            edge_seen.add(edge_key)
            edges.append(
                {
                    "from": source,
                    "to": destination,
                    "label": row["relation"],
                    "arrows": "to",
                    "direction": row["direction"],
                }
            )

        network_id = f"dep-network-{key_suffix}"
        zoom_in_id = f"{network_id}-zoom-in"
        zoom_out_id = f"{network_id}-zoom-out"
        fit_id = f"{network_id}-fit"
        incoming_id = f"{network_id}-filter-incoming"
        outgoing_id = f"{network_id}-filter-outgoing"
        class_id = f"{network_id}-filter-class"
        trigger_id = f"{network_id}-filter-trigger"
        object_id = f"{network_id}-filter-object"
        flow_id = f"{network_id}-filter-flow"
        metadata_id = f"{network_id}-filter-metadata"
        return f"""
<div class="graph-toolbar">
  <button id="{zoom_in_id}" type="button">Zoom +</button>
  <button id="{zoom_out_id}" type="button">Zoom -</button>
  <button id="{fit_id}" type="button">Centrer</button>
</div>
<div class="graph-filters">
  <label><input id="{incoming_id}" type="checkbox" checked>Afficher entrants</label>
  <label><input id="{outgoing_id}" type="checkbox" checked>Afficher sortants</label>
  <label><input id="{class_id}" type="checkbox" checked>Afficher classes</label>
  <label><input id="{trigger_id}" type="checkbox" checked>Afficher triggers</label>
  <label><input id="{object_id}" type="checkbox" checked>Afficher objets</label>
  <label><input id="{flow_id}" type="checkbox" checked>Afficher flows</label>
  <label><input id="{metadata_id}" type="checkbox" checked>Afficher metadata</label>
</div>
<div class="graph-legend">
  <span class="item"><span class="dot" style="background:#bfdbfe"></span>Apex/Trigger central</span>
  <span class="item"><span class="dot" style="background:#dbeafe"></span>Autres Apex/Trigger</span>
  <span class="item"><span class="dot" style="background:#dcfce7"></span>Objets</span>
  <span class="item"><span class="dot" style="background:#ffedd5"></span>Flows</span>
  <span class="item"><span class="dot" style="background:#f3e8ff"></span>Metadata</span>
</div>
<div id="{network_id}" class="dependency-graph"></div>
<script src="https://unpkg.com/vis-network@9.1.9/dist/vis-network.min.js"></script>
<script>
(() => {{
  const container = document.getElementById({json.dumps(network_id)});
  if (!container || typeof vis === "undefined") return;
  const centerId = {json.dumps(center_name)};
  const fullNodes = {json.dumps(list(nodes.values()))};
  const fullEdges = {json.dumps(edges)};
  const nodeMap = new Map(fullNodes.map((node) => [node.id, node]));
  const nodes = new vis.DataSet(fullNodes);
  const edges = new vis.DataSet(fullEdges);
  const network = new vis.Network(container, {{ nodes, edges }}, {{
    nodes: {{ borderWidth: 1, font: {{ face: "Arial", size: 13 }} }},
    edges: {{ color: "#64748b", arrows: "to", smooth: {{ type: "dynamic" }}, font: {{ align: "middle", size: 10 }} }},
    physics: {{ stabilization: true }},
    interaction: {{ hover: true, zoomView: true, dragView: true }}
  }});

  const zoomStep = 1.2;
  const zoomIn = document.getElementById({json.dumps(zoom_in_id)});
  const zoomOut = document.getElementById({json.dumps(zoom_out_id)});
  const fit = document.getElementById({json.dumps(fit_id)});
  const filterIncoming = document.getElementById({json.dumps(incoming_id)});
  const filterOutgoing = document.getElementById({json.dumps(outgoing_id)});
  const filterClass = document.getElementById({json.dumps(class_id)});
  const filterTrigger = document.getElementById({json.dumps(trigger_id)});
  const filterObject = document.getElementById({json.dumps(object_id)});
  const filterFlow = document.getElementById({json.dumps(flow_id)});
  const filterMetadata = document.getElementById({json.dumps(metadata_id)});

  const isKindEnabled = (kind) => {{
    const lowered = String(kind || "").toLowerCase();
    if (lowered === "class") return !filterClass || filterClass.checked;
    if (lowered === "trigger") return !filterTrigger || filterTrigger.checked;
    if (lowered === "apex") return (!filterClass || filterClass.checked) || (!filterTrigger || filterTrigger.checked);
    return true;
  }};
  const isCategoryEnabled = (category) => {{
    const lowered = String(category || "").toLowerCase();
    if (lowered === "objet") return !filterObject || filterObject.checked;
    if (lowered === "flow") return !filterFlow || filterFlow.checked;
    if (lowered === "metadata") return !filterMetadata || filterMetadata.checked;
    return true;
  }};

  const applyFilters = () => {{
    const allowIncoming = !filterIncoming || filterIncoming.checked;
    const allowOutgoing = !filterOutgoing || filterOutgoing.checked;
    const filteredEdges = [];
    const visibleNodeIds = new Set([centerId]);

    for (const edge of fullEdges) {{
      if (edge.direction === "Entrant" && !allowIncoming) continue;
      if (edge.direction === "Sortant" && !allowOutgoing) continue;

      const sourceNode = nodeMap.get(edge.from);
      const targetNode = nodeMap.get(edge.to);
      if (!sourceNode || !targetNode) continue;

      const linkedNode = edge.from === centerId ? targetNode : sourceNode;
      if (!isKindEnabled(linkedNode.componentKind)) continue;
      if (!isCategoryEnabled(linkedNode.category)) continue;

      filteredEdges.push(edge);
      visibleNodeIds.add(sourceNode.id);
      visibleNodeIds.add(targetNode.id);
    }}

    const filteredNodes = fullNodes.filter((node) => visibleNodeIds.has(node.id));
    nodes.clear();
    edges.clear();
    nodes.add(filteredNodes);
    edges.add(filteredEdges);
    network.fit({{ animation: false }});
  }};

  [filterIncoming, filterOutgoing, filterClass, filterTrigger, filterObject, filterFlow, filterMetadata].forEach((input) => {{
    if (input) input.addEventListener("change", applyFilters);
  }});

  if (zoomIn) {{
    zoomIn.addEventListener("click", () => {{
      const scale = network.getScale();
      network.moveTo({{ scale: scale * zoomStep }});
    }});
  }}
  if (zoomOut) {{
    zoomOut.addEventListener("click", () => {{
      const scale = network.getScale();
      network.moveTo({{ scale: scale / zoomStep }});
    }});
  }}
  if (fit) {{
    fit.addEventListener("click", () => network.fit({{ animation: true }}));
  }}
  network.on("doubleClick", (params) => {{
    const scale = network.getScale();
    const evt = params && params.event && params.event.srcEvent;
    const shift = !!(evt && evt.shiftKey);
    const factor = shift ? (1 / zoomStep) : zoomStep;
    network.moveTo({{ scale: scale * factor, animation: {{ duration: 150 }} }});
  }});
  network.once("stabilizationIterationsDone", () => {{
    network.setOptions({{ physics: false }});
  }});
  applyFilters();
}})();
</script>
"""

    def _build_flow_reference_index(
        self,
        flows: list[FlowInfo],
        flow_bodies: dict[str, str],
    ) -> dict[str, set[str]]:
        references: dict[str, set[str]] = {}
        patterns = {
            flow.name: re.compile(rf"\b{re.escape(flow.name)}\b", re.IGNORECASE)
            for flow in flows
        }
        for source in flows:
            body = flow_bodies.get(source.name, "")
            linked: set[str] = set()
            for target in flows:
                if target.name == source.name:
                    continue
                if patterns[target.name].search(body):
                    linked.add(target.name)
            references[source.name] = linked
        return references

    def _flow_dependencies(
        self,
        flow: FlowInfo,
        flow_ref_index: dict[str, set[str]],
        body: str,
        object_names: list[str],
        apex_names: list[str],
    ) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        seen: set[tuple[str, str, str, str]] = set()

        for target_name in sorted(flow_ref_index.get(flow.name, set()), key=str.lower):
            key = (target_name, "Sortant", "Reference flow", "Flow")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": target_name,
                    "category": "Flow",
                    "subtype": "Flow",
                    "direction": "Sortant",
                    "relation": "Reference flow",
                }
            )

        for source_name, targets in flow_ref_index.items():
            if source_name == flow.name or flow.name not in targets:
                continue
            key = (source_name, "Entrant", "Reference flow", "Flow")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": source_name,
                    "category": "Flow",
                    "subtype": "Flow",
                    "direction": "Entrant",
                    "relation": "Reference flow",
                }
            )

        if flow.start_object:
            key = (flow.start_object, "Sortant", "Objet de depart", "Objet")
            if key not in seen:
                seen.add(key)
                rows.append(
                    {
                        "name": flow.start_object,
                        "category": "Objet",
                        "subtype": "sObject",
                        "direction": "Sortant",
                        "relation": "Objet de depart",
                    }
                )

        for object_name in sorted(object_names, key=str.lower):
            if not re.search(rf"\b{re.escape(object_name)}\b", body):
                continue
            key = (object_name, "Sortant", "Usage objet", "Objet")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": object_name,
                    "category": "Objet",
                    "subtype": "sObject",
                    "direction": "Sortant",
                    "relation": "Usage objet",
                }
            )

        for apex_name in sorted(apex_names, key=str.lower):
            if not re.search(rf"\b{re.escape(apex_name)}\b", body):
                continue
            key = (apex_name, "Sortant", "Reference Apex", "Apex")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": apex_name,
                    "category": "Apex",
                    "subtype": "class",
                    "direction": "Sortant",
                    "relation": "Reference Apex",
                }
            )

        metadata_matches = sorted(set(re.findall(r"\b([A-Za-z_][A-Za-z0-9_]*__mdt)\b", body)))
        for metadata_name in metadata_matches:
            key = (metadata_name, "Sortant", "Reference metadata", "Metadata")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "name": metadata_name,
                    "category": "Metadata",
                    "subtype": "CustomMetadata",
                    "direction": "Sortant",
                    "relation": "Reference metadata",
                }
            )

        rows.sort(key=lambda item: (item["category"], item["direction"], item["name"].lower()))
        return rows

    def _dependency_node_color(self, category: str) -> dict[str, str]:
        palette = {
            "Apex": {"background": "#dbeafe", "border": "#3b82f6"},
            "Objet": {"background": "#dcfce7", "border": "#22c55e"},
            "Flow": {"background": "#ffedd5", "border": "#f97316"},
            "Metadata": {"background": "#f3e8ff", "border": "#a855f7"},
        }
        return palette.get(category, {"background": "#e2e8f0", "border": "#64748b"})

    def _tabbed_sections(self, group_id: str, sections: list[tuple[str, str]]) -> str:
        button_html: list[str] = []
        panel_html: list[str] = []
        used_slugs: set[str] = set()
        for index, (label, content) in enumerate(sections):
            base_slug = safe_slug(label) or f"tab-{index}"
            slug = base_slug
            suffix = 1
            while slug in used_slugs:
                suffix += 1
                slug = f"{base_slug}-{suffix}"
            used_slugs.add(slug)
            tab_id = f"{group_id}-tab-{slug}"
            panel_id = f"{group_id}-panel-{slug}"
            active_class = " active" if index == 0 else ""
            button_html.append(
                f"<button class='tab-button{active_class}' type='button' data-tab-group='{group_id}' data-tab-target='{panel_id}' id='{tab_id}'>{html_value(label)}</button>"
            )
            panel_html.append(
                f"<div class='tab-panel{active_class}' id='{panel_id}' data-tab-panel='{group_id}'>{content}</div>"
            )
        return (
            "<div class='tabs'>"
            f"<div class='tab-buttons'>{''.join(button_html)}</div>"
            f"{''.join(panel_html)}"
            "</div>"
        )

    def _page(self, title: str, body: str, current_path: Path, include_mermaid: bool = True) -> str:
        style_href = self._href(current_path, self.assets_dir / "style.css")
        if include_mermaid:
            mermaid_script = MERMAID_RUNTIME_SCRIPT
        else:
            mermaid_script = ""
        tabs_script = """
<script>
(() => {
  const renderMermaidIn = (panel) => {
    if (!panel) return;
    const fn = window.__renderMermaid;
    if (typeof fn === "function") {
      fn(panel);
      return;
    }
    if (!window.mermaid) return;
    const nodes = Array.from(panel.querySelectorAll('.mermaid:not([data-processed="true"])'));
    if (!nodes.length) return;
    try {
      window.mermaid.run({ nodes });
    } catch (err) {
      console.error("mermaid run (tab)", err);
    }
  };
  const activatePanel = (panel) => {
    if (!panel) return false;
    const group = panel.getAttribute("data-tab-panel");
    if (!group) return false;
    document.querySelectorAll(`[data-tab-group="${group}"]`).forEach((item) => item.classList.remove("active"));
    document.querySelectorAll(`[data-tab-panel="${group}"]`).forEach((item) => item.classList.remove("active"));
    panel.classList.add("active");
    const button = document.querySelector(`[data-tab-group="${group}"][data-tab-target="${panel.id}"]`);
    if (button) button.classList.add("active");
    renderMermaidIn(panel);
    return true;
  };
  document.querySelectorAll("[data-tab-group]").forEach((button) => {
    button.addEventListener("click", () => {
      const target = button.getAttribute("data-tab-target");
      if (!target) return;
      activatePanel(document.getElementById(target));
    });
  });
  const applyHash = () => {
    const hash = window.location.hash.slice(1);
    if (!hash) return;
    activatePanel(document.getElementById(hash));
  };
  window.addEventListener("hashchange", applyHash);
  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", applyHash);
  } else {
    applyHash();
  }
})();
</script>
""".strip()
        return f"""<!DOCTYPE html>
<html lang="fr">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html_value(title)}</title>
    <link rel="stylesheet" href="{style_href}">
    {mermaid_script}
  </head>
  <body>
    <div class="page">
      {body}
    </div>
    {tabs_script}
  </body>
</html>
"""

    def _href(self, from_path: Path, to_path: Path) -> str:
        return Path(os.path.relpath(to_path, from_path.parent)).as_posix()

    def _index_href(self, from_path: Path, tab_slug: str | None = None) -> str:
        href = self._href(from_path, self.output_dir / "index.html")
        if tab_slug:
            return f"{href}#index-panel-{tab_slug}"
        return href

    def _index_back_link(self, from_path: Path, tab_slug: str | None = None) -> str:
        href = self._index_href(from_path, tab_slug)
        return f"<div class=\"topnav\"><a href=\"{href}\">Retour a l'index</a></div>"
