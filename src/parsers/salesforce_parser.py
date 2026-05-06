from __future__ import annotations

import fnmatch
import json
import re
from collections import Counter
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

LogCallback = Callable[[str], None]

from src.core.models import (
    AgentInfo,
    ApexArtifact,
    CustomizationMetrics,
    FieldInfo,
    FieldPermission,
    FlowElementInfo,
    FlowInfo,
    GenAiPromptInfo,
    MetadataSnapshot,
    NamedAccess,
    ObjectInfo,
    ObjectPermission,
    RecordTypeInfo,
    RecordTypeVisibility,
    RelationshipInfo,
    SecurityArtifact,
    UserPermission,
    ValidationRuleInfo,
    VisibilityItem,
)
from src.core.utils import SF_NS, child_text, child_texts, parse_xml, to_bool


class SalesforceMetadataParser:
    """Parse a Salesforce DX source folder into a :class:`MetadataSnapshot`.

    Walks the well-known Salesforce metadata layout (objects, classes,
    triggers, flows, profiles, permission sets, etc.), produces structured
    Python dataclasses and applies an optional exclusion file so the caller
    can opt out of specific artefacts.
    """

    CATEGORY_ALIASES = {
        "all": "all",
        "global": "all",
        "objet": "object",
        "objets": "object",
        "object": "object",
        "objects": "object",
        "apex": "apex",
        "classe": "apex",
        "classes": "apex",
        "trigger": "apex",
        "triggers": "apex",
        "flow": "flow",
        "flows": "flow",
        "lwc": "lwc",
        "agent": "agent",
        "agents": "agent",
        "prompt": "prompt",
        "prompts": "prompt",
        "validation rule": "validation_rule",
        "validation rules": "validation_rule",
        "vr": "validation_rule",
        "omni": "omni",
        "omnistudio": "omni",
        "layout": "layout",
        "layouts": "layout",
        "flexipage": "flexipage",
        "flexipages": "flexipage",
        "lightning page": "flexipage",
        "lightning pages": "flexipage",
        "report": "report",
        "reports": "report",
        "dashboard": "dashboard",
        "dashboards": "dashboard",
        "profile": "profile",
        "profiles": "profile",
        "permission set": "permission_set",
        "permission sets": "permission_set",
        "permset": "permission_set",
        "permsets": "permission_set",
        "tab": "tab",
        "tabs": "tab",
        "application": "application",
        "applications": "application",
        "app": "application",
        "apps": "application",
        "ai_prediction": "ai_prediction",
        "ai_predictions": "ai_prediction",
        "business_rule": "business_rule",
        "business_rules": "business_rule",
        "bre": "business_rule",
    }

    def __init__(
        self,
        source_dir: str | Path,
        exclusion_config_path: str | Path | None = None,
        log_callback: LogCallback | None = None,
    ) -> None:
        self.source_dir = Path(source_dir).resolve()
        
        if exclusion_config_path:
            self.exclusion_config_path = Path(exclusion_config_path).resolve()
        else:
            # Default to exclusion.xlsx in the current directory if it exists
            candidate = Path("exclusion.xlsx")
            self.exclusion_config_path = candidate.resolve() if candidate.exists() else None

        self.log: LogCallback = log_callback or (lambda message: None)
        self.exclusion_rules: dict[str, list[str]] = self._load_exclusion_rules(
            self.exclusion_config_path
        )

    def parse(self) -> MetadataSnapshot:
        package_roots = self._resolve_package_roots()
        snapshot = MetadataSnapshot(source_dir=self.source_dir, package_roots=package_roots)

        objects: dict[str, ObjectInfo] = {}
        profiles: list[SecurityArtifact] = []
        permission_sets: list[SecurityArtifact] = []
        apex_artifacts: list[ApexArtifact] = []
        flows: list[FlowInfo] = []
        agents: list[AgentInfo] = []
        gen_ai_prompts: list[GenAiPromptInfo] = []
        metrics = CustomizationMetrics()

        for package_root in package_roots:
            self.log(f"Analyse du package {package_root}")
            objects_found = self._parse_objects(package_root)
            self.log(f"  - {len(objects_found)} objet(s) trouve(s)")
            objects.update(objects_found)
            
            profiles_found = self._parse_security_folder(package_root / "profiles", "profile")
            self.log(f"  - {len(profiles_found)} profil(s) trouve(s)")
            profiles.extend(profiles_found)
            
            permsets_found = self._parse_security_folder(package_root / "permissionsets", "permission_set")
            self.log(f"  - {len(permsets_found)} permission set(s) trouve(s)")
            permission_sets.extend(permsets_found)
            
            classes_found = self._parse_apex_folder(package_root / "classes", "class")
            self.log(f"  - {len(classes_found)} classe(s) Apex trouvee(s)")
            apex_artifacts.extend(classes_found)
            
            triggers_found = self._parse_apex_folder(package_root / "triggers", "trigger")
            self.log(f"  - {len(triggers_found)} trigger(s) Apex trouve(s)")
            apex_artifacts.extend(triggers_found)
            
            flows_found = self._parse_flows(package_root / "flows")
            self.log(f"  - {len(flows_found)} flow(s) trouve(s)")
            flows.extend(flows_found)

            agents_found = self._parse_agents(package_root / "agents")
            self.log(f"  - {len(agents_found)} agent(s) trouve(s)")
            agents.extend(agents_found)

            prompts_found = self._parse_gen_ai_prompts(package_root / "genAiPromptTemplates")
            self.log(f"  - {len(prompts_found)} prompt(s) trouve(s)")
            gen_ai_prompts.extend(prompts_found)

            metrics.lwc_count += len(
                [
                    path
                    for path in (package_root / "lwc").glob("*")
                    if path.is_dir() and not self._is_excluded("lwc", path.name)
                ]
            )
            metrics.flexipage_count += len(
                [
                    path
                    for path in (package_root / "flexipages").glob("*.flexipage-meta.xml")
                    if not self._is_excluded("flexipage", path.stem.replace(".flexipage-meta", ""))
                ]
            )
            metrics.layouts += len(
                [
                    path
                    for path in (package_root / "layouts").glob("*.layout-meta.xml")
                    if not self._is_excluded("layout", path.stem.replace(".layout-meta", ""))
                ]
            )
            metrics.custom_tabs += len(
                [
                    path
                    for path in (package_root / "tabs").glob("*.tab-meta.xml")
                    if "__" in path.stem and not self._is_excluded("tab", path.stem.replace(".tab-meta", ""))
                ]
            )
            metrics.custom_apps += len(
                [
                    path
                    for path in (package_root / "applications").glob("*.app-meta.xml")
                    if not path.stem.startswith("standard__") and not self._is_excluded("application", path.stem.replace(".app-meta", ""))
                ]
            )
            metrics.omni_scripts += len(
                [
                    path
                    for path in (package_root / "omniScripts").glob("*.os-meta.xml")
                    if not self._is_excluded("omni", path.stem.replace(".os-meta", ""))
                ]
            )
            metrics.omni_data_transforms += len(
                [
                    path
                    for path in (package_root / "omniDataTransforms").glob("*.rpt-meta.xml")
                    if not self._is_excluded("omni", path.stem.replace(".rpt-meta", ""))
                ]
            )
            metrics.einstein_predictions += len(
                [
                    path
                    for path in (package_root / "aiPredictions").glob("*.aiPrediction-meta.xml")
                    if not self._is_excluded("ai_prediction", path.stem.replace(".aiPrediction-meta", ""))
                ]
            )

        snapshot.objects = sorted(objects.values(), key=lambda item: item.api_name.lower())
        snapshot.profiles = sorted(profiles, key=lambda item: item.name.lower())
        snapshot.permission_sets = sorted(permission_sets, key=lambda item: item.name.lower())
        snapshot.apex_artifacts = sorted(apex_artifacts, key=lambda item: item.name.lower())
        snapshot.flows = sorted(flows, key=lambda item: item.name.lower())
        snapshot.agents = sorted(agents, key=lambda item: item.name.lower())
        snapshot.gen_ai_prompts = sorted(gen_ai_prompts, key=lambda item: item.name.lower())

        snapshot.profiles = [
            item
            for item in snapshot.profiles
            if not self._is_excluded("profile", item.name, item.label)
        ]
        snapshot.permission_sets = [
            item
            for item in snapshot.permission_sets
            if not self._is_excluded("permission_set", item.name, item.label)
        ]
        snapshot.objects = [
            item
            for item in snapshot.objects
            if not self._is_excluded("object", item.api_name, item.label)
        ]
        snapshot.apex_artifacts = [
            item
            for item in snapshot.apex_artifacts
            if not self._is_excluded("apex", item.name)
        ]
        snapshot.flows = [
            item
            for item in snapshot.flows
            if not self._is_excluded("flow", item.name, item.label)
        ]
        snapshot.agents = [
            item
            for item in snapshot.agents
            if not self._is_excluded("agent", item.name, item.label)
        ]
        snapshot.gen_ai_prompts = [
            item
            for item in snapshot.gen_ai_prompts
            if not self._is_excluded("prompt", item.name, item.label)
        ]
        snapshot.inventory = self._build_inventory(snapshot)

        metrics.custom_objects = sum(1 for item in snapshot.objects if item.custom)
        metrics.custom_fields = sum(1 for item in snapshot.objects for field in item.fields if field.custom)
        metrics.record_types = sum(len(item.record_types) for item in snapshot.objects)
        metrics.validation_rules = sum(len(item.validation_rules) for item in snapshot.objects)
        metrics.flows = len(snapshot.flows)
        metrics.apex_classes = sum(1 for item in snapshot.apex_artifacts if item.kind == "class")
        metrics.apex_triggers = sum(1 for item in snapshot.apex_artifacts if item.kind == "trigger")
        metrics.agents = len(snapshot.agents)
        metrics.gen_ai_prompts = len(snapshot.gen_ai_prompts)
        snapshot.metrics = metrics
        return snapshot

    def _load_exclusion_rules(
        self, config_path: Path | None
    ) -> dict[str, list[str]]:
        rules: dict[str, list[str]] = {
            val: [] for val in set(self.CATEGORY_ALIASES.values())
        }
        if "all" not in rules:
            rules["all"] = []
        
        if config_path is None:
            return rules
        if not config_path.exists():
            self.log(f"Fichier de configuration hors analyse introuvable: {config_path}")
            return rules

        workbook = load_workbook(config_path, data_only=True, read_only=True)
        sheet = None
        for candidate in workbook.sheetnames:
            if candidate.strip().lower() == "hors analyse":
                sheet = workbook[candidate]
                break
        if sheet is None:
            workbook.close()
            self.log("Onglet `hors analyse` introuvable dans le fichier de configuration.")
            return rules

        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if not values:
                continue
            if values[0].startswith("#"):
                continue

            category = self.CATEGORY_ALIASES.get(values[0].lower(), "")
            patterns: list[str] = []
            if category:
                raw = values[1] if len(values) > 1 else ""
                patterns = [part.strip() for part in re.split(r"[;,]", raw) if part.strip()]
                if not patterns and len(values) > 2:
                    patterns = [part.strip() for part in re.split(r"[;,]", values[2]) if part.strip()]
                if not patterns:
                    continue
            else:
                category = "all"
                patterns = [part.strip() for part in re.split(r"[;,]", values[0]) if part.strip()]

            for pattern in patterns:
                if pattern not in rules[category]:
                    rules[category].append(pattern)

        workbook.close()
        total = sum(len(items) for items in rules.values())
        if total:
            self.log(f"{total} regle(s) hors analyse chargee(s) depuis {config_path}.")
        return rules

    def _is_excluded(self, category: str, *names: str) -> bool:
        candidates = self.exclusion_rules.get(category, []) + self.exclusion_rules.get("all", [])
        if not candidates:
            return False
        targets = [name for name in names if name]
        if not targets:
            return False
        
        lowered_targets = [target.lower() for target in targets]
        normalized_targets = [self._normalize_exclusion_token(target) for target in targets]
        
        for pattern in candidates:
            lowered_pattern = pattern.lower()
            normalized_pattern = self._normalize_exclusion_token(pattern)
            
            for lowered_target, normalized_target in zip(lowered_targets, normalized_targets):
                # Exact match or glob match
                if fnmatch.fnmatch(lowered_target, lowered_pattern):
                    return True
                # Substring match (case insensitive)
                if lowered_pattern in lowered_target:
                    return True
                # Normalized match (removes spaces/underscores)
                if normalized_pattern and (normalized_pattern == normalized_target or normalized_pattern in normalized_target):
                    return True
        return False

    @staticmethod
    def _normalize_exclusion_token(value: str) -> str:
        return re.sub(r"[\s_]+", "", value or "").lower()

    def _build_inventory(self, snapshot: MetadataSnapshot) -> dict[str, list[dict[str, object]]]:
        return {
            "record_types": self._inventory_record_types(snapshot.objects),
            "layouts": self._inventory_layouts(snapshot.package_roots),
            "lightning_pages": self._inventory_flexipages(snapshot.package_roots),
            "validation_rules": self._inventory_validation_rules(snapshot.objects),
            "omnistudio": self._inventory_special_files(snapshot.package_roots, category="omnistudio"),
            "business_rules_engine": self._inventory_special_files(
                snapshot.package_roots, category="business_rules_engine"
            ),
            "flows": self._inventory_flows(snapshot.flows),
            "permission_sets": self._inventory_security(snapshot.permission_sets),
            "profiles": self._inventory_security(snapshot.profiles),
            "reports": self._inventory_reports(snapshot.package_roots),
            "dashboards": self._inventory_dashboards(snapshot.package_roots),
        }

    def _resolve_package_roots(self) -> list[Path]:
        config_path = self.source_dir / "sfdx-project.json"
        package_roots: list[Path] = []

        if config_path.exists():
            config = json.loads(config_path.read_text(encoding="utf-8"))
            for entry in config.get("packageDirectories", []):
                package_path = self.source_dir / entry.get("path", "")
                default_root = package_path / "main" / "default"
                if default_root.exists():
                    package_roots.append(default_root)
                elif package_path.exists():
                    package_roots.append(package_path)

        fallback = self.source_dir / "force-app" / "main" / "default"
        if fallback.exists() and fallback not in package_roots:
            package_roots.append(fallback)

        if not package_roots:
            self.log(f"Aucun packageDirectory trouve dans sfdx-project.json, utilisation de {self.source_dir}")
            package_roots.append(self.source_dir)

        return package_roots

    def _parse_objects(self, package_root: Path) -> dict[str, ObjectInfo]:
        objects_dir = package_root / "objects"
        parsed: dict[str, ObjectInfo] = {}
        if not objects_dir.exists():
            return parsed

        for object_dir in sorted(path for path in objects_dir.iterdir() if path.is_dir()):
            api_name = object_dir.name
            object_file = object_dir / f"{api_name}.object-meta.xml"
            info = ObjectInfo(api_name=api_name, custom="__" in api_name, source_path=object_file if object_file.exists() else object_dir)

            if object_file.exists():
                root = parse_xml(object_file)
                info.label = child_text(root, "label")
                info.plural_label = child_text(root, "pluralLabel")
                info.description = child_text(root, "description")
                info.deployment_status = child_text(root, "deploymentStatus")
                info.sharing_model = child_text(root, "sharingModel") or child_text(root, "externalSharingModel")
                info.visibility = child_text(root, "visibility")

            fields_dir = object_dir / "fields"
            if fields_dir.exists():
                for field_file in sorted(fields_dir.glob("*.field-meta.xml")):
                    info.fields.append(self._parse_field(field_file))

            record_types_dir = object_dir / "recordTypes"
            if record_types_dir.exists():
                for record_type_file in sorted(record_types_dir.glob("*.recordType-meta.xml")):
                    info.record_types.append(self._parse_record_type(record_type_file))

            validation_rules_dir = object_dir / "validationRules"
            if validation_rules_dir.exists():
                for validation_rule_file in sorted(validation_rules_dir.glob("*.validationRule-meta.xml")):
                    vr = self._parse_validation_rule(validation_rule_file)
                    if not self._is_excluded("validation_rule", f"{api_name}.{vr.full_name}", vr.full_name):
                        info.validation_rules.append(vr)

            info.relationships = [
                RelationshipInfo(
                    field_name=field.api_name,
                    relationship_type=field.data_type,
                    targets=field.reference_to,
                )
                for field in info.fields
                if field.reference_to
            ]
            parsed[api_name] = info

        return parsed

    def _parse_field(self, field_file: Path) -> FieldInfo:
        root = parse_xml(field_file)
        api_name = child_text(root, "fullName") or field_file.stem.replace(".field-meta", "")
        return FieldInfo(
            api_name=api_name,
            label=child_text(root, "label"),
            data_type=child_text(root, "type"),
            description=child_text(root, "description"),
            required=to_bool(child_text(root, "required")),
            custom="__" in api_name,
            reference_to=child_texts(root, "referenceTo"),
            relationship_name=child_text(root, "relationshipName"),
        )

    def _parse_record_type(self, record_type_file: Path) -> RecordTypeInfo:
        root = parse_xml(record_type_file)
        return RecordTypeInfo(
            full_name=child_text(root, "fullName") or record_type_file.stem.replace(".recordType-meta", ""),
            label=child_text(root, "label"),
            description=child_text(root, "description"),
            active=to_bool(child_text(root, "active")),
        )

    def _parse_validation_rule(self, validation_rule_file: Path) -> ValidationRuleInfo:
        root = parse_xml(validation_rule_file)
        return ValidationRuleInfo(
            full_name=child_text(root, "fullName")
            or validation_rule_file.stem.replace(".validationRule-meta", ""),
            active=to_bool(child_text(root, "active")),
            description=child_text(root, "description"),
            error_display_field=child_text(root, "errorDisplayField"),
            error_message=child_text(root, "errorMessage"),
            error_condition_formula=child_text(root, "errorConditionFormula"),
        )

    def _parse_security_folder(self, folder: Path, kind: str) -> list[SecurityArtifact]:
        artifacts: list[SecurityArtifact] = []
        if not folder.exists():
            return artifacts

        for meta_file in sorted(folder.glob("*.xml")):
            root = parse_xml(meta_file)
            artifact = SecurityArtifact(
                name=meta_file.name.split(".")[0],
                label=child_text(root, "label") or child_text(root, "fullName"),
                kind=kind,
                description=child_text(root, "description"),
                source_path=meta_file,
            )

            for node in root.findall("sf:objectPermissions", SF_NS):
                artifact.object_permissions.append(
                    ObjectPermission(
                        object_name=child_text(node, "object"),
                        allow_read=to_bool(child_text(node, "allowRead")),
                        allow_create=to_bool(child_text(node, "allowCreate")),
                        allow_edit=to_bool(child_text(node, "allowEdit")),
                        allow_delete=to_bool(child_text(node, "allowDelete")),
                        view_all_records=to_bool(child_text(node, "viewAllRecords")),
                        modify_all_records=to_bool(child_text(node, "modifyAllRecords")),
                    )
                )

            for node in root.findall("sf:fieldPermissions", SF_NS):
                artifact.field_permissions.append(
                    FieldPermission(
                        field_name=child_text(node, "field"),
                        readable=to_bool(child_text(node, "readable")),
                        editable=to_bool(child_text(node, "editable")),
                    )
                )

            for node in root.findall("sf:userPermissions", SF_NS):
                artifact.user_permissions.append(
                    UserPermission(
                        name=child_text(node, "name"),
                        enabled=to_bool(child_text(node, "enabled")),
                    )
                )

            for node in root.findall("sf:applicationVisibilities", SF_NS):
                artifact.application_visibilities.append(
                    VisibilityItem(
                        name=child_text(node, "application"),
                        visible=child_text(node, "visible"),
                        default=child_text(node, "default"),
                    )
                )

            for node in root.findall("sf:tabVisibilities", SF_NS) + root.findall("sf:tabSettings", SF_NS):
                artifact.tab_visibilities.append(
                    VisibilityItem(
                        name=child_text(node, "tab"),
                        visible=child_text(node, "visibility"),
                        default=child_text(node, "default"),
                    )
                )

            for node in root.findall("sf:classAccesses", SF_NS):
                artifact.class_accesses.append(
                    NamedAccess(
                        name=child_text(node, "apexClass"),
                        enabled=to_bool(child_text(node, "enabled")),
                    )
                )

            for node in root.findall("sf:flowAccesses", SF_NS):
                artifact.flow_accesses.append(
                    NamedAccess(
                        name=child_text(node, "flow"),
                        enabled=to_bool(child_text(node, "enabled")),
                    )
                )

            for node in root.findall("sf:pageAccesses", SF_NS):
                artifact.page_accesses.append(
                    NamedAccess(
                        name=child_text(node, "apexPage"),
                        enabled=to_bool(child_text(node, "enabled")),
                    )
                )

            for node in root.findall("sf:customPermissions", SF_NS):
                artifact.custom_permissions.append(
                    NamedAccess(
                        name=child_text(node, "name"),
                        enabled=to_bool(child_text(node, "enabled")),
                    )
                )

            for node in root.findall("sf:recordTypeVisibilities", SF_NS):
                artifact.record_type_visibilities.append(
                    RecordTypeVisibility(
                        record_type=child_text(node, "recordType"),
                        visible=to_bool(child_text(node, "visible")),
                        default=to_bool(child_text(node, "default")),
                    )
                )

            artifacts.append(artifact)

        return artifacts

    def _parse_apex_folder(self, folder: Path, kind: str) -> list[ApexArtifact]:
        artifacts: list[ApexArtifact] = []
        pattern = "*.cls" if kind == "class" else "*.trigger"
        if not folder.exists():
            return artifacts

        for source_file in sorted(folder.glob(pattern)):
            body = source_file.read_text(encoding="utf-8")
            meta_file = source_file.with_name(f"{source_file.name}-meta.xml")
            api_version = ""
            status = ""
            if meta_file.exists():
                root = parse_xml(meta_file)
                api_version = child_text(root, "apiVersion")
                status = child_text(root, "status")

            artifact = ApexArtifact(
                name=source_file.stem,
                kind=kind,
                body=body,
                source_path=source_file,
                api_version=api_version,
                status=status,
            )
            artifact.line_count = len(body.splitlines())
            artifact.method_count = len(
                re.findall(
                    r"(?mi)^\s*(?:public|private|protected|global)\s+(?:static\s+)?[\w<>\[\],]+\s+\w+\s*\(",
                    body,
                )
            )
            artifact.soql_count = len(re.findall(r"\[\s*SELECT\b|Database\.query\s*\(", body, re.IGNORECASE))
            artifact.sosl_count = len(re.findall(r"\[\s*FIND\b|Search\.query\s*\(", body, re.IGNORECASE))
            artifact.dml_count = len(
                re.findall(
                    r"(?i)\b(?:insert|update|upsert|delete|undelete|merge)\b|Database\.(?:insert|update|upsert|delete|undelete|merge)\s*\(",
                    body,
                )
            )
            artifact.comment_line_count = sum(
                1 for line in body.splitlines() if line.strip().startswith(("//", "/*", "*"))
            )
            artifact.system_debug_count = len(re.findall(r"System\.debug\s*\(", body))
            artifact.has_try_catch = "try" in body and "catch" in body
            sharing_match = re.search(
                r"(?i)\b(with sharing|without sharing|inherited sharing)\b", body
            )
            artifact.sharing_declaration = sharing_match.group(1) if sharing_match else ""
            artifact.is_test = bool(re.search(r"(?i)@isTest\b|\btestMethod\b", body))
            artifact.query_in_loop = bool(
                re.search(r"(?is)for\s*\(.*?\)\s*\{.{0,2000}?\[\s*SELECT\b", body)
            )
            artifact.dml_in_loop = bool(
                re.search(
                    r"(?is)for\s*\(.*?\)\s*\{.{0,2000}?\b(?:insert|update|upsert|delete|undelete|merge)\b",
                    body,
                )
            )
            artifacts.append(artifact)

        return artifacts

    def _parse_flows(self, folder: Path) -> list[FlowInfo]:
        flows: list[FlowInfo] = []
        if not folder.exists():
            return flows

        interesting_tags = [
            "actionCalls",
            "assignments",
            "collectionProcessors",
            "decisions",
            "formulas",
            "loops",
            "recordCreates",
            "recordDeletes",
            "recordLookups",
            "recordUpdates",
            "screens",
            "subflows",
            "transforms",
            "waits",
        ]

        for flow_file in sorted(folder.glob("*.flow-meta.xml")):
            root = parse_xml(flow_file)
            element_counts = Counter()
            elements: list[FlowElementInfo] = []
            described = 0
            undocumented = 0
            adjacency: dict[str, list[str]] = {}
            structural_types = {"decisions", "loops", "subflows"}
            nodes_by_name: dict[str, str] = {}

            for tag in interesting_tags:
                for node in root.findall(f"sf:{tag}", SF_NS):
                    element_counts[tag] += 1
                    description = child_text(node, "description")
                    if description:
                        described += 1
                    else:
                        undocumented += 1

                    name = child_text(node, "name")
                    if name:
                        nodes_by_name[name] = tag
                        adjacency.setdefault(name, [])

                    target = ""
                    connector = node.find("sf:connector/sf:targetReference", SF_NS)
                    if connector is not None and connector.text:
                        target = connector.text.strip()
                        if name:
                            adjacency[name].append(target)

                    if tag == "decisions":
                        for rule in node.findall("sf:rules", SF_NS):
                            rule_connector = rule.find("sf:connector", SF_NS)
                            rule_target = (
                                child_text(rule_connector, "targetReference") if rule_connector is not None else ""
                            )
                            if name and rule_target:
                                adjacency[name].append(rule_target)
                        default_connector = node.find("sf:defaultConnector", SF_NS)
                        default_target = (
                            child_text(default_connector, "targetReference")
                            if default_connector is not None
                            else ""
                        )
                        if name and default_target:
                            adjacency[name].append(default_target)
                    elif tag == "loops":
                        next_connector = node.find("sf:nextValueConnector", SF_NS)
                        next_target = (
                            child_text(next_connector, "targetReference") if next_connector is not None else ""
                        )
                        if name and next_target:
                            adjacency[name].append(next_target)
                        end_connector = node.find("sf:noMoreValuesConnector", SF_NS)
                        end_target = (
                            child_text(end_connector, "targetReference") if end_connector is not None else ""
                        )
                        if name and end_target:
                            adjacency[name].append(end_target)

                    fault_connector = node.find("sf:faultConnector", SF_NS)
                    fault_target = (
                        child_text(fault_connector, "targetReference") if fault_connector is not None else ""
                    )
                    if name and fault_target:
                        adjacency[name].append(fault_target)

                    elements.append(
                        FlowElementInfo(
                            element_type=tag,
                            name=name,
                            label=child_text(node, "label"),
                            description=description,
                            target=target,
                        )
                    )

            variables = root.findall("sf:variables", SF_NS)
            variable_total = len(variables)
            variable_input = 0
            variable_output = 0
            for variable in variables:
                if to_bool(child_text(variable, "isInput")):
                    variable_input += 1
                if to_bool(child_text(variable, "isOutput")):
                    variable_output += 1

            start_node = ""
            start = root.find("sf:start", SF_NS)
            if start is not None:
                start_connector = start.find("sf:connector", SF_NS)
                start_node = child_text(start_connector, "targetReference") if start_connector is not None else ""

            max_width = 1
            for decision in root.findall("sf:decisions", SF_NS):
                width = len(decision.findall("sf:rules", SF_NS))
                if decision.find("sf:defaultConnector", SF_NS) is not None:
                    width += 1
                max_width = max(max_width, width)

            min_height = 0
            max_height = 0
            max_depth = 0
            if start_node and start_node in nodes_by_name:
                paths = self._flow_paths(start_node, adjacency)
                if paths:
                    min_height = min(len(path) for path in paths)
                    max_height = max(len(path) for path in paths)
                    for path in paths:
                        depth = sum(
                            1
                            for node_name in path
                            if nodes_by_name.get(node_name) in structural_types
                        )
                        max_depth = max(max_depth, depth)

            flow = FlowInfo(
                name=flow_file.stem.replace(".flow-meta", ""),
                label=child_text(root, "label"),
                description=child_text(root, "description"),
                process_type=child_text(root, "processType"),
                status=child_text(root, "status"),
                api_version=child_text(root, "apiVersion"),
                trigger_type=child_text(root.find("sf:start", SF_NS), "triggerType")
                if root.find("sf:start", SF_NS) is not None
                else "",
                start_object=child_text(root.find("sf:start", SF_NS), "object")
                if root.find("sf:start", SF_NS) is not None
                else "",
                source_path=flow_file,
                element_counts=dict(element_counts),
                described_elements=described,
                undocumented_elements=undocumented,
                total_elements=sum(element_counts.values()),
                variable_total=variable_total,
                variable_input=variable_input,
                variable_output=variable_output,
                max_width=max_width,
                min_height=min_height,
                max_height=max_height,
                max_depth=max_depth,
                elements=elements,
            )
            flows.append(flow)

        return flows

    def _flow_paths(self, start_node: str, adjacency: dict[str, list[str]]) -> list[list[str]]:
        paths: list[list[str]] = []
        stack: list[tuple[str, list[str]]] = [(start_node, [start_node])]
        safeguard = 0

        while stack and safeguard < 5000:
            safeguard += 1
            current, path = stack.pop()
            neighbors = adjacency.get(current, [])
            if not neighbors:
                paths.append(path)
                continue

            advanced = False
            for neighbor in neighbors:
                if neighbor and neighbor not in path:
                    stack.append((neighbor, [*path, neighbor]))
                    advanced = True
            if not advanced:
                paths.append(path)

        return paths

    def _inventory_record_types(self, objects: list[ObjectInfo]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for item in objects:
            for record_type in item.record_types:
                rows.append(
                    {
                        "Objet": item.api_name,
                        "Record Type": record_type.full_name,
                        "Label": record_type.label,
                        "Actif": record_type.active,
                        "Description": record_type.description,
                    }
                )
        return rows

    def _inventory_validation_rules(self, objects: list[ObjectInfo]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for item in objects:
            for rule in item.validation_rules:
                rows.append(
                    {
                        "Objet": item.api_name,
                        "Regle": rule.full_name,
                        "Active": rule.active,
                        "Description": rule.description,
                        "ChampErreur": rule.error_display_field,
                    }
                )
        return rows

    def _inventory_security(self, artifacts: list[SecurityArtifact]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for artifact in artifacts:
            rows.append(
                {
                    "Nom": artifact.name,
                    "Label": artifact.label,
                    "Description": artifact.description,
                    "DroitsObjet": len(artifact.object_permissions),
                    "DroitsChamp": len(artifact.field_permissions),
                    "PermissionsSysteme": len(artifact.user_permissions),
                    "Applications": len(artifact.application_visibilities),
                    "Flows": len(artifact.flow_accesses),
                    "RecordTypes": len(artifact.record_type_visibilities),
                    "Source": self._safe_relative_path(artifact.source_path),
                }
            )
        return rows

    def _inventory_flows(self, flows: list[FlowInfo]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for flow in flows:
            rows.append(
                {
                    "Nom": flow.name,
                    "Label": flow.label,
                    "Type": flow.process_type,
                    "Statut": flow.status,
                    "Objet": flow.start_object,
                    "Declencheur": flow.trigger_type,
                    "Complexite": flow.complexity_level,
                    "Score": flow.complexity_score,
                    "Elements": flow.total_elements,
                    "Source": self._safe_relative_path(flow.source_path),
                }
            )
        return rows

    def _inventory_layouts(self, package_roots: list[Path]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for package_root in package_roots:
            folder = package_root / "layouts"
            if not folder.exists():
                continue
            for meta_file in sorted(folder.glob("*.layout-meta.xml")):
                name = meta_file.stem.replace(".layout-meta", "")
                if self._is_excluded("layout", name):
                    continue
                root = parse_xml(meta_file)
                rows.append(
                    {
                        "Objet": meta_file.stem.split("-")[0],
                        "Layout": meta_file.stem.replace(".layout-meta", ""),
                        "Sections": len(root.findall("sf:layoutSections", SF_NS)),
                        "RelatedLists": len(root.findall("sf:relatedLists", SF_NS)),
                        "BoutonsExclus": len(root.findall("sf:excludeButtons", SF_NS)),
                        "MiniLayout": root.find("sf:miniLayout", SF_NS) is not None,
                        "Source": self._safe_relative_path(meta_file),
                    }
                )
        return rows

    def _inventory_flexipages(self, package_roots: list[Path]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for package_root in package_roots:
            folder = package_root / "flexipages"
            if not folder.exists():
                continue
            for meta_file in sorted(folder.glob("*.flexipage-meta.xml")):
                name = meta_file.stem.replace(".flexipage-meta", "")
                if self._is_excluded("flexipage", name):
                    continue
                root = parse_xml(meta_file)
                rows.append(
                    {
                        "NomAPI": meta_file.stem.replace(".flexipage-meta", ""),
                        "Label": child_text(root, "masterLabel"),
                        "Type": child_text(root, "type"),
                        "Objet": child_text(root, "sobjectType"),
                        "Template": child_text(root.find("sf:template", SF_NS), "name")
                        if root.find("sf:template", SF_NS) is not None
                        else "",
                        "Regions": len(root.findall("sf:flexiPageRegions", SF_NS)),
                        "Composants": len(root.findall(".//sf:componentInstance", SF_NS)),
                        "Source": self._safe_relative_path(meta_file),
                    }
                )
        return rows

    def _inventory_reports(self, package_roots: list[Path]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for package_root in package_roots:
            folder = package_root / "reports"
            if not folder.exists():
                continue
            for meta_file in sorted(folder.rglob("*.report-meta.xml")):
                name = meta_file.stem.replace(".report-meta", "")
                if self._is_excluded("report", name):
                    continue
                root = parse_xml(meta_file)
                relative = meta_file.relative_to(folder)
                rows.append(
                    {
                        "Dossier": str(relative.parent).replace("\\", "/") if str(relative.parent) != "." else "",
                        "Nom": meta_file.stem.replace(".report-meta", ""),
                        "Label": child_text(root, "name") or child_text(root, "fullName"),
                        "Description": child_text(root, "description"),
                        "TypeRapport": child_text(root, "reportType"),
                        "Filtres": len(root.findall("sf:filter", SF_NS)) + len(root.findall("sf:standardFilter", SF_NS)),
                        "Colonnes": len(root.findall("sf:columns", SF_NS)),
                        "Source": self._safe_relative_path(meta_file),
                    }
                )
        return rows

    def _inventory_dashboards(self, package_roots: list[Path]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for package_root in package_roots:
            folder = package_root / "dashboards"
            if not folder.exists():
                continue
            for meta_file in sorted(folder.rglob("*.dashboard-meta.xml")):
                name = meta_file.stem.replace(".dashboard-meta", "")
                if self._is_excluded("dashboard", name):
                    continue
                root = parse_xml(meta_file)
                relative = meta_file.relative_to(folder)
                rows.append(
                    {
                        "Dossier": str(relative.parent).replace("\\", "/") if str(relative.parent) != "." else "",
                        "Nom": meta_file.stem.replace(".dashboard-meta", ""),
                        "Titre": child_text(root, "title"),
                        "Type": child_text(root, "dashboardType"),
                        "RunningUser": child_text(root, "runningUser"),
                        "Composants": len(root.findall("sf:dashboardGridComponents", SF_NS)),
                        "Source": self._safe_relative_path(meta_file),
                    }
                )
        return rows

    def _inventory_special_files(
        self, package_roots: list[Path], category: str
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        if category == "omnistudio":
            keywords = ("omni", "omnistudio", "datasource", "vlocity")
            known_folders = {
                "omniscripts",
                "omniuicards",
                "omnidatatransforms",
                "omniprocesses",
                "omnistudio",
            }
            label = "OmniStudio"
        else:
            keywords = (
                "decision",
                "expression",
                "calculationmatrix",
                "ruleset",
                "recommendationstrategy",
            )
            known_folders = {
                "decisionmatrices",
                "decisionmatrixdefinitions",
                "decisiontables",
                "expressionsets",
                "expressionsetdefinitions",
                "calculationmatrices",
                "recommendationstrategies",
            }
            label = "Business Rules Engine"

        for package_root in package_roots:
            for meta_file in sorted(package_root.rglob("*")):
                if meta_file.is_dir() or meta_file.suffix.lower() != ".xml":
                    continue

                folder_name = meta_file.parent.name.lower()
                file_name = meta_file.name.lower()
                stem = meta_file.stem.lower()
                name = meta_file.stem.split(".")[0]
                
                if self._is_excluded("omni" if category == "omnistudio" else "business_rule", name):
                    continue

                if folder_name in known_folders or any(token in file_name or token in stem for token in keywords):
                    rows.append(
                        {
                            "Nom": meta_file.stem.split(".")[0],
                            "Categorie": label,
                            "Dossier": meta_file.parent.name,
                            "TypeFichier": "".join(meta_file.suffixes),
                            "Source": self._safe_relative_path(meta_file),
                        }
                    )

        unique_rows: list[dict[str, object]] = []
        seen_sources: set[str] = set()
        for row in rows:
            source = str(row["Source"])
            if source in seen_sources:
                continue
            seen_sources.add(source)
            unique_rows.append(row)
        return unique_rows

    def _parse_agents(self, folder: Path) -> list[AgentInfo]:
        agents: list[AgentInfo] = []
        if not folder.exists():
            return agents
        for agent_file in sorted(folder.glob("*.agent-meta.xml")):
            root = parse_xml(agent_file)
            agents.append(
                AgentInfo(
                    name=agent_file.stem.replace(".agent-meta", ""),
                    label=child_text(root, "label"),
                    description=child_text(root, "description"),
                    source_path=agent_file,
                )
            )
        return agents

    def _parse_gen_ai_prompts(self, folder: Path) -> list[GenAiPromptInfo]:
        prompts: list[GenAiPromptInfo] = []
        if not folder.exists():
            return prompts
        for prompt_file in sorted(folder.glob("*.genAiPromptTemplate-meta.xml")):
            root = parse_xml(prompt_file)
            prompts.append(
                GenAiPromptInfo(
                    name=prompt_file.stem.replace(".genAiPromptTemplate-meta", ""),
                    label=child_text(root, "masterLabel"),
                    description=child_text(root, "description"),
                    source_path=prompt_file,
                )
            )
        return prompts

    def _safe_relative_path(self, path: Path) -> str:
        try:
            return str(path.relative_to(self.source_dir)).replace("\\", "/")
        except ValueError:
            return str(path)
