import pandas as pd
from openpyxl import load_workbook

try:
    file_path = "exclusion.xlsx"
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet_name = None
    for candidate in workbook.sheetnames:
        if candidate.strip().lower() == "hors analyse":
            sheet_name = candidate
            break
    
    if sheet_name:
        sheet = workbook[sheet_name]
        print(f"Sheet found: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            print(row)
    else:
        print("Sheet 'hors analyse' not found.")
        print(f"Available sheets: {workbook.sheetnames}")
    workbook.close()
except Exception as e:
    print(f"Error: {e}")
